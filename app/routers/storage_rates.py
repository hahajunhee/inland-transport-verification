from io import BytesIO
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.services import storage_rate_service
from app import data_store

router = APIRouter()


# ─── Pydantic 스키마 ─────────────────────────────────────────────────

class StorageRateCreate(BaseModel):
    odcy_name:           Optional[str]   = ""
    odcy_terminal_type:  Optional[str]   = ""
    odcy_location:       Optional[str]   = ""
    dest_port_type:      Optional[str]   = ""
    dest_terminal_type:  Optional[str]   = ""
    storage_tier1:  Optional[float] = None
    storage_tier2:  Optional[float] = None
    storage_tier3:  Optional[float] = None
    storage_tier4:  Optional[float] = None
    storage_tier5:  Optional[float] = None
    storage_tier6:  Optional[float] = None
    handling_tier1: Optional[float] = None
    handling_tier2: Optional[float] = None
    handling_tier3: Optional[float] = None
    handling_tier4: Optional[float] = None
    handling_tier5: Optional[float] = None
    handling_tier6: Optional[float] = None
    shuttle_tier1:  Optional[float] = None
    shuttle_tier2:  Optional[float] = None
    shuttle_tier3:  Optional[float] = None
    shuttle_tier4:  Optional[float] = None
    shuttle_tier5:  Optional[float] = None
    shuttle_tier6:  Optional[float] = None
    memo:           Optional[str]   = ""


# ─── CRUD ─────────────────────────────────────────────────────────

@router.get("/")
def list_storage_rates():
    return storage_rate_service.get_all_storage_rates()


@router.post("/", status_code=201)
def add_storage_rate(body: StorageRateCreate):
    return storage_rate_service.create_storage_rate(body.model_dump())


@router.put("/{rate_id}")
def edit_storage_rate(rate_id: int, body: StorageRateCreate):
    obj = storage_rate_service.update_storage_rate(rate_id, body.model_dump())
    if not obj:
        raise HTTPException(status_code=404, detail="요율을 찾을 수 없습니다.")
    return obj


@router.delete("/{rate_id}", status_code=204)
def remove_storage_rate(rate_id: int):
    if not storage_rate_service.delete_storage_rate(rate_id):
        raise HTTPException(status_code=404, detail="요율을 찾을 수 없습니다.")


# ─── 엑셀 템플릿 다운로드 ─────────────────────────────────────────────

def _style_header(ws, headers, col_widths):
    ws.append(headers)
    fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


@router.get("/template")
def download_template():
    rates = storage_rate_service.get_all_storage_rates()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "보관료_상하차료 요율"
    _style_header(ws, ["ODCY명", "odcy터미널구분", "ODCY_위치", "도착지포트구분", "도착지터미널구분",
                       "보관료 단가", "상하차료 단가", "비고"],
                  [20, 20, 20, 20, 20, 14, 14, 30])
    for r in rates:
        ws.append([r.get("odcy_name", ""), r.get("odcy_terminal_type", ""),
                   r.get("odcy_location", ""), r.get("dest_port_type", ""),
                   r.get("dest_terminal_type", ""),
                   r.get("storage_tier1"), r.get("handling_tier1"), r.get("memo", "")])
    if not rates:
        ws.append(["세방(주)", "배후단지", "부산신항", "부산북항", "", 10000, 8000, "예시 (등록 후 삭제)"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''storage_rates_template.xlsx"},
    )


# ─── 엑셀 업로드 (전체 교체) ──────────────────────────────────────────

@router.post("/upload")
async def upload_storage_rates(file: UploadFile = File(...)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, detail=f"올바른 xlsx 파일이 아닙니다: {e}")

    sheet_name = None
    for sn in wb.sheetnames:
        if "보관료" in sn or "상하차" in sn or "요율" in sn:
            sheet_name = sn
            break
    if sheet_name is None and wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    if sheet_name is None:
        raise HTTPException(400, detail="처리할 시트가 없습니다.")

    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

    if "보관료 단가" not in col_map and "상하차료 단가" not in col_map:
        raise HTTPException(400, detail="헤더에 '보관료 단가' 또는 '상하차료 단가' 컬럼이 없습니다.")

    def to_float(v):
        if v is None or str(v).strip() == "":
            return None
        try:
            return float(v)
        except Exception:
            return None

    data_store.save("storage_rates.json", [])
    success, failed, new_items, next_id = 0, [], [], 1

    # 헤더에 [OM-A] 등 접미사가 붙어있을 수 있으므로 부분 매칭 헬퍼
    def _find_col(*keywords):
        """col_map에서 keyword를 포함하는 헤더를 찾아 인덱스 반환."""
        for kw in keywords:
            # 정확한 매칭 우선
            if kw in col_map:
                return col_map[kw]
        for kw in keywords:
            # 부분 매칭
            for header_name, idx in col_map.items():
                if kw in header_name:
                    return idx
        return None

    om_a_col     = _find_col("ODCY도착지명")                     # OM-A
    odcy_col     = _find_col("ODCY명")                            # OM-B
    term_col     = _find_col("odcy터미널구분", "터미널구분", "단지구분")  # OM-C
    loc_col      = _find_col("ODCY_위치")                         # OM-D
    dpt_col      = _find_col("도착지포트구분")
    dtt_col      = _find_col("도착지터미널구분")
    storage_col  = _find_col("보관료 단가")
    handling_col = _find_col("상하차료 단가")
    memo_col     = _find_col("비고")

    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        def gv(col_idx):
            return row[col_idx].value if col_idx is not None else None

        om_a               = str(gv(om_a_col) or "").strip()
        odcy_name          = str(gv(odcy_col) or "").strip()
        odcy_terminal_type = str(gv(term_col) or "").strip()
        odcy_location      = str(gv(loc_col) or "").strip()
        dest_port_type     = str(gv(dpt_col) or "").strip()
        dest_terminal_type = str(gv(dtt_col) or "").strip()
        storage_unit       = to_float(gv(storage_col))
        handling_unit      = to_float(gv(handling_col))

        if not odcy_name and not odcy_terminal_type and storage_unit is None and handling_unit is None:
            continue

        new_items.append({
            "id": next_id,
            "om_a": om_a,
            "odcy_name": odcy_name,
            "odcy_terminal_type": odcy_terminal_type,
            "odcy_location": odcy_location,
            "dest_port_type": dest_port_type,
            "dest_terminal_type": dest_terminal_type,
            "storage_tier1": storage_unit,
            "handling_tier1": handling_unit,
            "memo": str(gv(memo_col) or "").strip(),
        })
        next_id += 1
        success += 1

    data_store.save("storage_rates.json", new_items)
    return {"success": success, "failed": failed}
