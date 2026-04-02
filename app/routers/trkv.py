from io import BytesIO
from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.services import trkv_service, storage_rate_service
from app import data_store

router = APIRouter()


# ─── Pydantic 스키마 ─────────────────────────────────────────────────

class PortMappingCreate(BaseModel):
    excel_name:    str
    port_type:     str
    terminal_type: Optional[str] = ""


class OdcyMappingCreate(BaseModel):
    odcy_destination_name: str
    odcy_name:             str
    odcy_terminal_type:    Optional[str] = ""
    odcy_location:         Optional[str] = ""


class DepartureMappingCreate(BaseModel):
    departure_name: str
    departure_code: str


class RouteCreate(BaseModel):
    pickup_port: str
    departure_code: str
    dest_port: str
    tier1: Optional[float] = None
    tier2: Optional[float] = None
    tier3: Optional[float] = None
    tier4: Optional[float] = None
    tier5: Optional[float] = None
    tier6: Optional[float] = None
    memo: Optional[str] = None


class ContainerTierItem(BaseModel):
    cont_type: str
    is_dg: bool
    tier_number: Optional[int] = None


class ContainerTierBulk(BaseModel):
    items: list[ContainerTierItem]


# ─── 포트명 매핑 CRUD ─────────────────────────────────────────────────

@router.get("/port-mappings")
def list_port_mappings():
    return trkv_service.get_all_port_mappings()


@router.post("/port-mappings", status_code=201)
def add_port_mapping(body: PortMappingCreate):
    try:
        return trkv_service.create_port_mapping(body.excel_name, body.port_type, body.terminal_type or "")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/port-mappings/{mapping_id}")
def edit_port_mapping(mapping_id: int, body: PortMappingCreate):
    obj = trkv_service.update_port_mapping(mapping_id, body.excel_name, body.port_type, body.terminal_type or "")
    if not obj:
        raise HTTPException(status_code=404, detail="포트 매핑을 찾을 수 없습니다.")
    return obj


@router.delete("/port-mappings/{mapping_id}", status_code=204)
def remove_port_mapping(mapping_id: int):
    if not trkv_service.delete_port_mapping(mapping_id):
        raise HTTPException(status_code=404, detail="포트 매핑을 찾을 수 없습니다.")


# ─── 출하지 매핑 CRUD ─────────────────────────────────────────────────

@router.get("/departure-mappings")
def list_departure_mappings():
    return trkv_service.get_all_departure_mappings()


@router.post("/departure-mappings", status_code=201)
def add_departure_mapping(body: DepartureMappingCreate):
    try:
        return trkv_service.create_departure_mapping(body.departure_name, body.departure_code)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/departure-mappings/{mapping_id}")
def edit_departure_mapping(mapping_id: int, body: DepartureMappingCreate):
    obj = trkv_service.update_departure_mapping(mapping_id, body.departure_name, body.departure_code)
    if not obj:
        raise HTTPException(status_code=404, detail="출하지 매핑을 찾을 수 없습니다.")
    return obj


@router.delete("/departure-mappings/{mapping_id}", status_code=204)
def remove_departure_mapping(mapping_id: int):
    if not trkv_service.delete_departure_mapping(mapping_id):
        raise HTTPException(status_code=404, detail="출하지 매핑을 찾을 수 없습니다.")


# ─── ODCY 매핑 CRUD ──────────────────────────────────────────────────

@router.get("/odcy-mappings")
def list_odcy_mappings():
    return trkv_service.get_all_odcy_mappings()


@router.post("/odcy-mappings", status_code=201)
def add_odcy_mapping(body: OdcyMappingCreate):
    try:
        return trkv_service.create_odcy_mapping(
            body.odcy_destination_name, body.odcy_name,
            body.odcy_terminal_type or "", body.odcy_location or "",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/odcy-mappings/{mapping_id}")
def edit_odcy_mapping(mapping_id: int, body: OdcyMappingCreate):
    obj = trkv_service.update_odcy_mapping(
        mapping_id, body.odcy_destination_name, body.odcy_name,
        body.odcy_terminal_type or "", body.odcy_location or "",
    )
    if not obj:
        raise HTTPException(status_code=404, detail="ODCY 매핑을 찾을 수 없습니다.")
    return obj


@router.delete("/odcy-mappings/{mapping_id}", status_code=204)
def remove_odcy_mapping(mapping_id: int):
    if not trkv_service.delete_odcy_mapping(mapping_id):
        raise HTTPException(status_code=404, detail="ODCY 매핑을 찾을 수 없습니다.")


# ─── 구간요율 CRUD ─────────────────────────────────────────────────────

@router.get("/routes")
def list_routes():
    return trkv_service.get_all_routes()


@router.post("/routes", status_code=201)
def add_route(body: RouteCreate):
    return trkv_service.create_route(body.model_dump())


@router.put("/routes/{route_id}")
def edit_route(route_id: int, body: RouteCreate):
    obj = trkv_service.update_route(route_id, body.model_dump())
    if not obj:
        raise HTTPException(status_code=404, detail="구간 요율을 찾을 수 없습니다.")
    return obj


@router.delete("/routes/{route_id}", status_code=204)
def remove_route(route_id: int):
    if not trkv_service.delete_route(route_id):
        raise HTTPException(status_code=404, detail="구간 요율을 찾을 수 없습니다.")


# ─── 컨테이너 티어 (TRKV) ───────────────────────────────────────────

@router.get("/container-tiers")
def list_container_tiers():
    return trkv_service.get_all_container_tiers()


@router.post("/container-tiers/bulk")
def save_container_tiers(body: ContainerTierBulk):
    return trkv_service.bulk_save_container_tiers([i.model_dump() for i in body.items])


@router.put("/container-tiers/{tier_id}")
def edit_container_tier(tier_id: int, tier_number: Optional[int] = None):
    obj = trkv_service.update_container_tier(tier_id, tier_number)
    if not obj:
        raise HTTPException(status_code=404, detail="컨테이너 티어를 찾을 수 없습니다.")
    return obj


# ─── 컨테이너 티어 (보관료/상하차료/셔틀비) ──────────────────────────

@router.get("/storage-container-tiers")
def list_storage_container_tiers():
    return trkv_service.get_all_storage_container_tiers()


@router.post("/storage-container-tiers/bulk")
def save_storage_container_tiers(body: ContainerTierBulk):
    return trkv_service.bulk_save_storage_container_tiers([i.model_dump() for i in body.items])


# ─── 헬퍼 ─────────────────────────────────────────────────────────────

def _style_header(ws, headers: list, col_widths: list):
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def _find_col(col_map: dict, *names) -> Optional[int]:
    """여러 후보 열제목 중 처음으로 매칭되는 열 인덱스 반환 (0-based)."""
    for name in names:
        if name in col_map:
            return col_map[name]
    return None


# ─── 통합 엑셀 템플릿 (현재 등록된 데이터 포함) ───────────────────────

@router.get("/template")
def download_unified_template():
    """현재 등록된 데이터를 포함한 통합 양식 다운로드 (전체 교체용, 5-시트)"""
    port_mappings      = trkv_service.get_all_port_mappings()
    departure_mappings = trkv_service.get_all_departure_mappings()
    odcy_mappings      = trkv_service.get_all_odcy_mappings()
    routes             = trkv_service.get_all_routes()
    storage_rates      = storage_rate_service.get_all_storage_rates()

    wb = openpyxl.Workbook()

    # ── Sheet 1: 포트명 매핑 ─────────────────────────────────────────
    ws_pm = wb.active
    ws_pm.title = "포트명 매핑"
    _style_header(
        ws_pm,
        ["포트명(엑셀원본명) [PM-A]", "포트 구분 [PM-B]", "터미널구분 [PM-C]"],
        [34, 18, 22],
    )
    for pm in port_mappings:
        ws_pm.append([pm["excel_name"], pm["port_type"],
                      pm.get("terminal_type") or pm.get("zone_type", "")])
    if not port_mappings:
        ws_pm.append(["부산신항BPTS", "부산신항", ""])
        ws_pm.append(["북컨배후단지", "부산북항", ""])

    # ── Sheet 2: 출하지 매핑 ─────────────────────────────────────────
    ws_dm = wb.create_sheet("출하지 매핑")
    _style_header(
        ws_dm,
        ["출하지명(엑셀원본명) [DM-A]", "출하지코드 [DM-B]"],
        [34, 18],
    )
    for dm in departure_mappings:
        ws_dm.append([dm["departure_name"], dm["departure_code"]])
    if not departure_mappings:
        ws_dm.append(["아산공장", "AS"])
        ws_dm.append(["울산출하지", "UL"])

    # ── Sheet 3: ODCY 매핑 ───────────────────────────────────────────
    ws_om = wb.create_sheet("ODCY 매핑")
    _style_header(
        ws_om,
        ["ODCY도착지명(엑셀원본명) [OM-A]", "ODCY명 [OM-B]", "odcy터미널구분 [OM-C]", "ODCY_위치 [OM-D]"],
        [38, 22, 22, 22],
    )
    for om in odcy_mappings:
        ws_om.append([
            om["odcy_destination_name"], om["odcy_name"],
            om.get("odcy_terminal_type") or om.get("terminal_type", ""),
            om.get("odcy_location", ""),
        ])
    if not odcy_mappings:
        ws_om.append(["SB청암", "세방(주)", "배후단지", "부산신항"])

    # ── Sheet 4: TRKV 구간 요율 ──────────────────────────────────────
    ws_rt = wb.create_sheet("TRKV 구간 요율")
    _style_header(
        ws_rt,
        ["픽업항", "출하지코드", "도착항", "티어1", "티어2", "티어3", "티어4", "티어5", "티어6", "비고"],
        [15, 15, 15, 12, 12, 12, 12, 12, 12, 25],
    )
    for r in routes:
        ws_rt.append([
            r.get("pickup_port"),
            r.get("departure_code", r.get("departure_name")),
            r.get("dest_port"),
            r.get("tier1"), r.get("tier2"), r.get("tier3"),
            r.get("tier4"), r.get("tier5"), r.get("tier6"),
            r.get("memo"),
        ])
    if not routes:
        ws_rt.append(["부산신항", "AS", "부산북항", 100000, 110000, 120000, 130000, 140000, 150000, "예시 (등록 후 삭제)"])

    # ── Sheet 5: 보관료_상하차료_셔틀비 요율 ──────────────────────────
    # A열: ODCY도착지명(엑셀원본명) [OM-A]  ← ODCY 매핑의 원본명
    # B열: 포트명(엑셀원본명) [PM-A]         ← 포트명 매핑의 원본명
    # 두 매핑의 모든 조합을 미리 생성하고 기존 요율이 있으면 채워 넣음
    ws_sr = wb.create_sheet("보관료_상하차료_셔틀비 요율")
    sr_headers = [
        "ODCY도착지명(엑셀원본명) [OM-A]", "포트명(엑셀원본명) [PM-A]",
        "보관료_T1", "보관료_T2", "보관료_T3", "보관료_T4", "보관료_T5", "보관료_T6",
        "상하차료_T1", "상하차료_T2", "상하차료_T3", "상하차료_T4", "상하차료_T5", "상하차료_T6",
        "셔틀비_T1", "셔틀비_T2", "셔틀비_T3", "셔틀비_T4", "셔틀비_T5", "셔틀비_T6",
        "비고",
    ]
    _style_header(ws_sr, sr_headers, [38, 34] + [11] * 18 + [25])

    # 기존 요율 룩업: (odcy_name, odcy_terminal_type, odcy_location, dest_port_type, dest_terminal_type) → rate
    sr_lookup: dict = {}
    for sr in storage_rates:
        key = (
            sr.get("odcy_name", ""), sr.get("odcy_terminal_type", ""),
            sr.get("odcy_location", ""), sr.get("dest_port_type", ""),
            sr.get("dest_terminal_type", ""),
        )
        sr_lookup[key] = sr

    # ODCY 매핑 / 포트명 매핑 룩업
    odcy_by_dest = {m["odcy_destination_name"]: m for m in odcy_mappings}
    port_by_excel = {m["excel_name"]: m for m in port_mappings}

    rows_written = 0
    if odcy_mappings and port_mappings:
        for om in odcy_mappings:
            odcy_dest = om["odcy_destination_name"]
            for pm in port_mappings:
                port_excel = pm["excel_name"]
                # 5키로 기존 요율 조회
                key = (
                    om.get("odcy_name", ""), om.get("odcy_terminal_type", ""),
                    om.get("odcy_location", ""), pm.get("port_type", ""),
                    pm.get("terminal_type", ""),
                )
                sr = sr_lookup.get(key, {})
                ws_sr.append([
                    odcy_dest, port_excel,
                    sr.get("storage_tier1"), sr.get("storage_tier2"), sr.get("storage_tier3"),
                    sr.get("storage_tier4"), sr.get("storage_tier5"), sr.get("storage_tier6"),
                    sr.get("handling_tier1"), sr.get("handling_tier2"), sr.get("handling_tier3"),
                    sr.get("handling_tier4"), sr.get("handling_tier5"), sr.get("handling_tier6"),
                    sr.get("shuttle_tier1"), sr.get("shuttle_tier2"), sr.get("shuttle_tier3"),
                    sr.get("shuttle_tier4"), sr.get("shuttle_tier5"), sr.get("shuttle_tier6"),
                    sr.get("memo", ""),
                ])
                rows_written += 1

    if rows_written == 0:
        # 매핑 데이터 없을 때 예시 행
        ws_sr.append(["SB청암", "부산신항BPTS",
                      10000, 11000, 12000, None, None, None,
                      8000,  9000,  10000, None, None, None,
                      5000,  6000,  7000,  None, None, None,
                      "예시 (등록 후 삭제)"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''unified_template.xlsx"},
    )


# ─── 통합 업로드 (항상 전체 교체) ────────────────────────────────────

@router.post("/upload")
async def upload_unified(file: UploadFile = File(...)):
    """통합 업로드: 내용이 있는 시트를 전체 교체로 처리"""
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, detail=f"올바른 xlsx 파일이 아닙니다: {e}")

    try:
        return _process_upload(wb)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, detail=f"업로드 처리 중 오류가 발생했습니다: {e}")


def _process_upload(wb):

    def to_float(v):
        if v is None or str(v).strip() == "":
            return None
        try:
            return float(v)
        except Exception:
            return None

    def has_data(ws):
        return any(
            any(c.value is not None for c in row)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row or 1)
        )

    result = {}

    # ── 포트명 매핑 시트 ─────────────────────────────────────────────
    if "포트명 매핑" in wb.sheetnames:
        ws = wb["포트명 매핑"]
        header  = [cell.value for cell in ws[1]]
        col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

        excel_col = _find_col(col_map,
            "포트명(엑셀원본명) [PM-A]", "포트명(엑셀원본명)", "엑셀 원본명")
        port_col  = _find_col(col_map,
            "포트 구분 [PM-B]", "포트 구분")
        term_col  = _find_col(col_map,
            "터미널구분 [PM-C]", "터미널구분", "단지구분")

        if has_data(ws) and excel_col is not None and port_col is not None:
            data_store.save("port_mappings.json", [])
            success, failed, new_items, next_id = 0, [], [], 1
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                excel_name    = row[excel_col].value
                port_type     = row[port_col].value
                terminal_type = row[term_col].value if term_col is not None else None
                if not excel_name and not port_type:
                    continue
                if not excel_name or not port_type:
                    failed.append({"row": i, "error": "포트명(엑셀원본명), 포트 구분 모두 필수입니다."})
                    continue
                new_items.append({
                    "id": next_id,
                    "excel_name":    str(excel_name).strip(),
                    "port_type":     str(port_type).strip(),
                    "terminal_type": str(terminal_type).strip() if terminal_type else "",
                })
                next_id += 1; success += 1
            data_store.save("port_mappings.json", new_items)
            result["포트명 매핑"] = {"success": success, "failed": failed}

    # ── 출하지 매핑 시트 ─────────────────────────────────────────────
    if "출하지 매핑" in wb.sheetnames:
        ws = wb["출하지 매핑"]
        header  = [cell.value for cell in ws[1]]
        col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

        dep_name_col = _find_col(col_map,
            "출하지명(엑셀원본명) [DM-A]", "출하지명(엑셀원본명)",
            "출하지명 (엑셀 원본명)", "출하지명")
        dep_code_col = _find_col(col_map,
            "출하지코드 [DM-B]", "출하지코드")

        if has_data(ws) and dep_name_col is not None and dep_code_col is not None:
            data_store.save("departure_mappings.json", [])
            success, failed, new_items, next_id = 0, [], [], 1
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                dep_name = row[dep_name_col].value
                dep_code = row[dep_code_col].value
                if not dep_name and not dep_code:
                    continue
                if not dep_name or not dep_code:
                    failed.append({"row": i, "error": "출하지명, 출하지코드 모두 필수입니다."})
                    continue
                new_items.append({"id": next_id,
                                  "departure_name": str(dep_name).strip(),
                                  "departure_code": str(dep_code).strip()})
                next_id += 1; success += 1
            data_store.save("departure_mappings.json", new_items)
            result["출하지 매핑"] = {"success": success, "failed": failed}

    # ── ODCY 매핑 시트 ───────────────────────────────────────────────
    if "ODCY 매핑" in wb.sheetnames:
        ws = wb["ODCY 매핑"]
        header  = [cell.value for cell in ws[1]]
        col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}

        dest_col    = _find_col(col_map,
            "ODCY도착지명(엑셀원본명) [OM-A]", "ODCY도착지명(엑셀원본명)",
            "ODCY 도착지명 (엑셀 원본명)", "ODCY 도착지명")
        name_col    = _find_col(col_map, "ODCY명 [OM-B]", "ODCY명")
        term_col_om = _find_col(col_map,
            "odcy터미널구분 [OM-C]", "odcy터미널구분", "터미널구분")
        loc_col     = _find_col(col_map, "ODCY_위치 [OM-D]", "ODCY_위치")

        if has_data(ws) and dest_col is not None and name_col is not None:
            data_store.save("odcy_mappings.json", [])
            success, failed, new_items, next_id = 0, [], [], 1
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                odcy_dest = row[dest_col].value
                odcy_name = row[name_col].value
                odcy_term = row[term_col_om].value if term_col_om is not None else None
                odcy_loc  = row[loc_col].value if loc_col is not None else None
                if not odcy_dest and not odcy_name:
                    continue
                if not odcy_dest or not odcy_name:
                    failed.append({"row": i, "error": "ODCY 도착지명, ODCY명 모두 필수입니다."})
                    continue
                new_items.append({
                    "id": next_id,
                    "odcy_destination_name": str(odcy_dest).strip(),
                    "odcy_name": str(odcy_name).strip(),
                    "odcy_terminal_type": str(odcy_term).strip() if odcy_term else "",
                    "odcy_location": str(odcy_loc).strip() if odcy_loc else "",
                })
                next_id += 1; success += 1
            data_store.save("odcy_mappings.json", new_items)
            result["ODCY 매핑"] = {"success": success, "failed": failed}

    # ── TRKV 구간 요율 시트 ──────────────────────────────────────────
    if "TRKV 구간 요율" in wb.sheetnames:
        ws = wb["TRKV 구간 요율"]
        header  = [cell.value for cell in ws[1]]
        col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
        dep_col_name = "출하지코드" if "출하지코드" in col_map else "출하지명"
        if has_data(ws) and all(c in col_map for c in ["픽업항", dep_col_name, "도착항"]):
            data_store.save("trkv_routes.json", [])

            def gv(row, name):
                idx = col_map.get(name)
                return row[idx].value if idx is not None else None

            success, failed, new_routes, next_id = 0, [], [], 1
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                pickup  = gv(row, "픽업항")
                dep_val = gv(row, dep_col_name)
                dest    = gv(row, "도착항")
                if not pickup and not dep_val and not dest:
                    continue
                if not pickup or not dep_val or not dest:
                    failed.append({"row": i, "error": "픽업항, 출하지코드, 도착항은 필수입니다."})
                    continue
                data = {
                    "id": next_id,
                    "pickup_port":    str(pickup).strip(),
                    "departure_code": str(dep_val).strip(),
                    "dest_port":      str(dest).strip(),
                    "tier1": to_float(gv(row, "티어1")),
                    "tier2": to_float(gv(row, "티어2")),
                    "tier3": to_float(gv(row, "티어3")),
                    "tier4": to_float(gv(row, "티어4")),
                    "tier5": to_float(gv(row, "티어5")),
                    "tier6": to_float(gv(row, "티어6")),
                    "memo": str(gv(row, "비고") or "").strip() or None,
                }
                new_routes.append(data)
                next_id += 1; success += 1
            data_store.save("trkv_routes.json", new_routes)
            result["TRKV 구간 요율"] = {"success": success, "failed": failed}

    # ── 보관료_상하차료_셔틀비 요율 시트 ─────────────────────────────
    for sheet_name in wb.sheetnames:
        if "보관료" in sheet_name or "상하차" in sheet_name or "셔틀" in sheet_name:
            ws = wb[sheet_name]
            header  = [cell.value for cell in ws[1]]
            col_map = {str(v).strip(): i for i, v in enumerate(header) if v is not None}
            tier_keys = [k for k in col_map if "_T" in k or "보관료 단가" in k or "상하차료 단가" in k]
            if not (has_data(ws) and tier_keys):
                break

            data_store.save("storage_rates.json", [])

            # 새 형식: A=ODCY도착지명(엑셀원본명), B=포트명(엑셀원본명) → 매핑으로 5키 resolve
            odcy_dest_col = _find_col(col_map,
                "ODCY도착지명(엑셀원본명) [OM-A]", "ODCY도착지명(엑셀원본명)")
            port_excel_col = _find_col(col_map,
                "포트명(엑셀원본명) [PM-A]", "포트명(엑셀원본명)")

            use_new_format = (odcy_dest_col is not None and port_excel_col is not None)

            if use_new_format:
                # 매핑 룩업 테이블 (방금 저장된 JSON에서 로드)
                _odcy_map = {
                    m["odcy_destination_name"]: m
                    for m in data_store.load("odcy_mappings.json")
                }
                _port_map = {
                    m["excel_name"]: m
                    for m in data_store.load("port_mappings.json")
                }

            else:
                # 구 형식: ODCY명 직접 입력
                odcy_col  = _find_col(col_map, "ODCY명")
                term_col2 = _find_col(col_map,
                    "odcy터미널구분", "터미널구분", "단지구분")
                loc_col2  = _find_col(col_map, "ODCY_위치")
                dpt_col   = _find_col(col_map, "도착지포트구분")
                dtt_col   = _find_col(col_map, "도착지터미널구분")

            memo_col = _find_col(col_map, "비고")

            success, failed, new_items, next_id = 0, [], [], 1
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                def _gv2(col_idx, _row=row):
                    return _row[col_idx].value if col_idx is not None else None

                if use_new_format:
                    odcy_dest_val  = str(_gv2(odcy_dest_col) or "").strip()
                    port_excel_val = str(_gv2(port_excel_col) or "").strip()

                    odcy_entry = _odcy_map.get(odcy_dest_val, {})
                    port_entry = _port_map.get(port_excel_val, {})

                    odcy_name          = odcy_entry.get("odcy_name", "")
                    odcy_terminal_type = odcy_entry.get("odcy_terminal_type", "")
                    odcy_location      = odcy_entry.get("odcy_location", "")
                    dest_port_type     = port_entry.get("port_type", "")
                    dest_terminal_type = port_entry.get("terminal_type", "")
                else:
                    odcy_name          = str(_gv2(odcy_col) or "").strip()
                    odcy_terminal_type = str(_gv2(term_col2) or "").strip()
                    odcy_location      = str(_gv2(loc_col2) or "").strip()
                    dest_port_type     = str(_gv2(dpt_col) or "").strip()
                    dest_terminal_type = str(_gv2(dtt_col) or "").strip()

                sr_cols = {
                    "보관료_T": "storage_tier",
                    "상하차료_T": "handling_tier",
                    "셔틀비_T": "shuttle_tier",
                }
                obj = {
                    "id": next_id,
                    "odcy_name": odcy_name,
                    "odcy_terminal_type": odcy_terminal_type,
                    "odcy_location": odcy_location,
                    "dest_port_type": dest_port_type,
                    "dest_terminal_type": dest_terminal_type,
                    "memo": str(_gv2(memo_col) or "").strip(),
                }
                all_none = True
                for prefix, key in sr_cols.items():
                    for t in range(1, 7):
                        col_key = f"{prefix}{t}"
                        val = to_float(_gv2(col_map.get(col_key)))
                        obj[f"{key}{t}"] = val
                        if val is not None:
                            all_none = False
                # 구버전 단가 컬럼 fallback
                if "보관료 단가" in col_map:
                    v = to_float(_gv2(col_map["보관료 단가"]))
                    obj["storage_tier1"] = v; all_none = all_none and (v is None)
                if "상하차료 단가" in col_map:
                    v = to_float(_gv2(col_map["상하차료 단가"]))
                    obj["handling_tier1"] = v; all_none = all_none and (v is None)

                # 새 형식: 두 키 모두 빈 값이면 스킵
                if use_new_format:
                    if not odcy_dest_val and not port_excel_val:
                        continue
                else:
                    if not odcy_name and not odcy_terminal_type and all_none:
                        continue

                new_items.append(obj)
                next_id += 1; success += 1

            data_store.save("storage_rates.json", new_items)
            result["보관료_상하차료_셔틀비 요율"] = {"success": success, "failed": failed}
            break  # 시트 하나만 처리

    if not result:
        raise HTTPException(400, detail="처리할 수 있는 시트가 없습니다. 통합 양식을 사용하세요.")

    return {"sheets": result}
