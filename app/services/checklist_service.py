"""
1단계 체크리스트 검증 서비스
7개 Stage에 걸쳐 검증파일의 데이터 정합성을 확인합니다.
"""
import json
import pandas as pd
from io import BytesIO
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from app import data_store

# ─── 체크리스트 전용 컬럼 매핑 ────────────────────────────────────────
CHECKLIST_COLUMN_MAP = {
    "D/G여부":              "dg_flag",
    "Cont.Category":        "cont_category",
    "Cont Size":            "cont_size",
    "Cont.Type":            "cont_type",
    "픽업지":               "pickup_code",
    "픽업지명":             "pickup_name",
    "상세 ODCY":            "odcy_code",
    "상세 ODCY명":          "odcy_name",
    "ODCY도착지명":         "odcy_destination_name",
    "ODCY 도착지명":        "odcy_destination_name",
    "출하지명":             "departure_name",
    "도착지":               "dest_code",
    "도착지명":             "dest_name",
    "출하일":               "transport_date",
    "Contrainer No.":       "container_no",
    "Container No.":        "container_no",
    "C/Invoice No.":        "c_invoice_no",
    "FWO Doc.":             "fwo_doc",
    "Quantity":             "quantity",
    "Weekend / Holiday":    "weekend_holiday",
    "ODCY 반입일":          "odcy_in_date",
    "ODCY 반출일":          "odcy_out_date",
    # 체크리스트 전용 추가 컬럼
    "ODCY":                 "odcy_col",
    "터미널 반입일":        "terminal_in_date",
    "선적일":               "shipping_date",
    "Frozen Container":     "frozen_container",
    "Waiting (ODCY)":       "waiting_odcy",
    "운송유형":             "transport_type",
    "운송모드":             "transport_mode",
    "터미널명":             "terminal_name",
}

# ─── 데이터 저장 경로 ──────────────────────────────────────────────────
CHECKLIST_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "checklist"

# ─── 유저 정의 휴일 ───────────────────────────────────────────────────
HOLIDAYS_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "user_holidays.json"


def load_user_holidays() -> list[str]:
    if not HOLIDAYS_PATH.exists():
        return []
    try:
        return json.loads(HOLIDAYS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_user_holidays(dates: list[str]):
    HOLIDAYS_PATH.parent.mkdir(parents=True, exist_ok=True)
    HOLIDAYS_PATH.write_text(json.dumps(sorted(set(dates)), ensure_ascii=False), encoding="utf-8")


def _is_holiday_or_sunday(d: date) -> tuple[bool, str]:
    if d.weekday() == 6:
        return True, "일요일"
    ds = d.isoformat()
    if ds in _holiday_set:
        return True, "공휴일(사용자 지정)"
    return False, ""


_holiday_set: set[str] = set()


def _refresh_holiday_set():
    global _holiday_set
    _holiday_set = set(load_user_holidays())


# ─── 파싱 ──────────────────────────────────────────────────────────────

def _parse_date_val(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s in ("nan", "None", "NaT"):
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def parse_checklist_excel(file_bytes: bytes) -> list[dict]:
    df = pd.read_excel(BytesIO(file_bytes), header=None, dtype=str)

    header_row_idx = None
    for i, row in df.iterrows():
        row_values = [str(v) for v in row.values]
        if any(col in row_values for col in ["Cont.Category", "픽업지", "D/G여부"]):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("헤더 행을 찾을 수 없습니다. (Cont.Category, 픽업지 등)")

    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)

    available = {col: mapped for col, mapped in CHECKLIST_COLUMN_MAP.items() if col in df.columns}

    rows = []
    for idx, raw_row in df.iterrows():
        row: dict = {}
        for excel_col, internal_key in available.items():
            val = raw_row.get(excel_col)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                row[internal_key] = None
            else:
                s = str(val).strip()
                row[internal_key] = s if s not in ("nan", "None", "NaT", "") else None

        pickup = row.get("pickup_code")
        if not pickup:
            continue

        row["row_number"] = int(idx) + int(header_row_idx) + 1
        rows.append(row)

    return rows


# ─── 공통 유틸 ─────────────────────────────────────────────────────────

def _is_blank(val) -> bool:
    return val is None or str(val).strip() == ""


def _is_not_blank(val) -> bool:
    return not _is_blank(val)


# ─── Stage 1: 필수 항목 공란 체크 ──────────────────────────────────────

def check_stage1(rows: list[dict]) -> list[dict]:
    always_check = {
        "pickup_name": "픽업지명",
        "departure_name": "출하지명",
        "dest_name": "도착지명",
        "transport_date": "출하일",
        "terminal_in_date": "터미널 반입일",
        "shipping_date": "선적일",
    }
    odcy_cols = {
        "odcy_code": "상세 ODCY",
        "odcy_in_date": "ODCY 반입일",
        "odcy_out_date": "ODCY 반출일",
        "odcy_destination_name": "ODCY 도착지명",
    }

    errors = []
    for row in rows:
        is_port_drt = (row.get("odcy_col") or "").strip().upper() == "PORT_DRT"

        for field, col_name in always_check.items():
            if _is_blank(row.get(field)):
                errors.append({
                    "row_number": row["row_number"],
                    "container_no": row.get("container_no"),
                    "column": col_name,
                    "value": "",
                    "reason": f"{col_name}이(가) 공란입니다",
                })

        if is_port_drt:
            for field, col_name in odcy_cols.items():
                if _is_not_blank(row.get(field)):
                    errors.append({
                        "row_number": row["row_number"],
                        "container_no": row.get("container_no"),
                        "column": col_name,
                        "value": row.get(field),
                        "reason": f"직반입(PORT_DRT)인데 {col_name}에 값이 입력되어 있습니다",
                    })
        else:
            for field, col_name in odcy_cols.items():
                if _is_blank(row.get(field)):
                    errors.append({
                        "row_number": row["row_number"],
                        "container_no": row.get("container_no"),
                        "column": col_name,
                        "value": "",
                        "reason": f"{col_name}이(가) 공란입니다 (ODCY가 PORT_DRT가 아닌 경우 필수)",
                    })

    return errors


# ─── Stage 2: 리퍼/DG 체크 ────────────────────────────────────────────

def check_stage2(rows: list[dict]) -> list[dict]:
    errors = []
    for row in rows:
        category = (row.get("cont_category") or "").strip().upper()
        frozen = (row.get("frozen_container") or "").strip().upper()

        if category == "RE" and frozen != "X":
            errors.append({
                "row_number": row["row_number"],
                "container_no": row.get("container_no"),
                "column": "Frozen Container",
                "value": frozen or "(공란)",
                "reason": "리퍼(RE) 컨테이너인데 Frozen Container가 체크(X)되지 않았습니다",
            })

        dg_val = row.get("dg_flag")
        if _is_not_blank(dg_val):
            errors.append({
                "row_number": row["row_number"],
                "container_no": row.get("container_no"),
                "column": "D/G여부",
                "value": dg_val,
                "reason": "D/G여부 열은 공란이어야 합니다",
            })

    return errors


# ─── Stage 3: 공휴일/주말 체크 ────────────────────────────────────────

def check_stage3(rows: list[dict]) -> list[dict]:
    errors = []
    for row in rows:
        d = _parse_date_val(row.get("transport_date"))
        if d is None:
            continue

        wh = (row.get("weekend_holiday") or "").strip().upper()
        is_checked = wh == "X"
        is_hol, day_type = _is_holiday_or_sunday(d)

        if is_hol and not is_checked:
            errors.append({
                "row_number": row["row_number"],
                "container_no": row.get("container_no"),
                "column": "Weekend / Holiday",
                "value": "(미체크)",
                "reason": f"출하일({d})이 {day_type}인데 Weekend/Holiday가 체크되지 않았습니다",
            })
        elif not is_hol and is_checked:
            errors.append({
                "row_number": row["row_number"],
                "container_no": row.get("container_no"),
                "column": "Weekend / Holiday",
                "value": "X",
                "reason": f"출하일({d})이 평일인데 Weekend/Holiday가 체크(X)되어 있습니다",
            })

    return errors


# ─── Stage 4: 직상차 체크 ─────────────────────────────────────────────

def check_stage4(rows: list[dict], container_numbers: list[str]) -> list[dict]:
    cn_set = {cn.strip().upper() for cn in container_numbers if cn.strip()}

    cn_map: dict[str, list[dict]] = {}
    for row in rows:
        cn = (row.get("container_no") or "").strip().upper()
        if cn:
            cn_map.setdefault(cn, []).append(row)

    errors = []

    # 1) 입력한 컨테이너 → Waiting (ODCY)가 X가 아니면 오류
    for cn in cn_set:
        if cn not in cn_map:
            errors.append({
                "row_number": None,
                "container_no": cn,
                "column": "Container No.",
                "value": "(미존재)",
                "reason": f"컨테이너 {cn}이(가) 검증파일에 존재하지 않습니다",
            })
            continue

        for row in cn_map[cn]:
            waiting = (row.get("waiting_odcy") or "").strip().upper()
            if waiting != "X":
                errors.append({
                    "row_number": row["row_number"],
                    "container_no": cn,
                    "column": "Waiting (ODCY)",
                    "value": waiting or "(공란)",
                    "reason": "직상차 컨테이너인데 Waiting (ODCY)가 'X'가 아닙니다",
                })

    # 2) 입력하지 않은 컨테이너인데 Waiting (ODCY)가 X인 행 → 오류
    for row in rows:
        cn = (row.get("container_no") or "").strip().upper()
        if not cn or cn in cn_set:
            continue
        waiting = (row.get("waiting_odcy") or "").strip().upper()
        if waiting == "X":
            errors.append({
                "row_number": row["row_number"],
                "container_no": cn,
                "column": "Waiting (ODCY)",
                "value": "X",
                "reason": "직상차 목록에 없는 컨테이너인데 Waiting (ODCY)가 'X'로 체크되어 있습니다",
            })

    return errors


# ─── Stage 5: 운송유형/모드 누락 체크 ─────────────────────────────────

def check_stage5(rows: list[dict]) -> list[dict]:
    errors = []
    checks = [("transport_type", "운송유형"), ("transport_mode", "운송모드")]
    for row in rows:
        for field, col_name in checks:
            if _is_blank(row.get(field)):
                errors.append({
                    "row_number": row["row_number"],
                    "container_no": row.get("container_no"),
                    "column": col_name,
                    "value": "",
                    "reason": f"{col_name}이(가) 공란입니다",
                })
    return errors


# ─── Stage 6: 왕복/편도 체크 ──────────────────────────────────────────

def _resolve_port_type(name: str) -> tuple[Optional[str], bool]:
    """포트명매핑: excel_name(PM-A) → port_type(PM-B). 반환: (PM-B, found)"""
    if not name:
        return None, False
    name = name.strip()
    items = data_store.load("port_mappings.json")
    for m in items:
        if m.get("excel_name") == name:
            return m.get("port_type"), True
    return None, False


def _resolve_om_d(odcy_dest_name: str) -> Optional[str]:
    """ODCY매핑: odcy_destination_name(OM-A) → odcy_location(OM-D)"""
    if not odcy_dest_name:
        return None
    name = odcy_dest_name.strip()
    if not name:
        return None
    items = data_store.load("odcy_mappings.json")
    for m in items:
        if m.get("odcy_destination_name") == name:
            loc = m.get("odcy_location") or ""
            return loc.strip() if loc.strip() else None
    return None


_SINPORT_OMD = {"KRPUSN", "PUSN16", "PUSN7"}


def _om_d_to_port(om_d: str) -> str:
    """OM-D 코드 → 포트구분 (KRPUSN/PUSN16/PUSN7→부산신항, 그 외→부산북항)"""
    if om_d and om_d.strip() in _SINPORT_OMD:
        return "부산신항"
    return "부산북항"


def check_stage6(rows: list[dict]) -> list[dict]:
    errors = []
    for row in rows:
        pickup_name = row.get("pickup_name")
        odcy_dest_name = row.get("odcy_destination_name")
        dest_name = row.get("dest_name")
        transport_type = (row.get("transport_type") or "").strip()

        if not pickup_name:
            continue

        pickup_zone, p_found = _resolve_port_type(pickup_name)
        if not p_found:
            continue

        has_odcy_dest = odcy_dest_name and str(odcy_dest_name).strip() != ""

        if has_odcy_dest:
            # ODCY도착지명 있음 → OM-A매칭 → OM-D → 포트변환 후 비교
            om_d = _resolve_om_d(odcy_dest_name)
            if om_d is None:
                continue  # ODCY매핑에 없으면 스킵
            dest_port = _om_d_to_port(om_d)
            expected = "왕복" if pickup_zone == dest_port else "편도"
            detail = (
                f"픽업지({pickup_name})→PM-B:{pickup_zone}, "
                f"ODCY도착지명({odcy_dest_name})→OM-D:{om_d}→{dest_port}"
            )
        else:
            # ODCY도착지명 공란 → 기존 방식: 도착지명 PM-B 비교
            if not dest_name:
                continue
            dest_zone, d_found = _resolve_port_type(dest_name)
            if not d_found:
                continue
            expected = "왕복" if pickup_zone == dest_zone else "편도"
            detail = (
                f"픽업지({pickup_name})→PM-B:{pickup_zone}, "
                f"도착지({dest_name})→PM-B:{dest_zone}"
            )

        if transport_type != expected:
            errors.append({
                "row_number": row["row_number"],
                "container_no": row.get("container_no"),
                "column": "운송유형",
                "value": transport_type or "(공란)",
                "reason": f"{detail} → '{expected}'이어야 하는데 '{transport_type}'입니다",
            })

    return errors


# ─── Stage 7: 일자 순서 체크 ──────────────────────────────────────────

_DATE_FIELDS = [
    ("transport_date", "출하일"),
    ("odcy_in_date", "ODCY 반입일"),
    ("odcy_out_date", "ODCY 반출일"),
    ("terminal_in_date", "터미널 반입일"),
    ("shipping_date", "선적일"),
]


def check_stage7(rows: list[dict]) -> list[dict]:
    errors = []
    for row in rows:
        dates = []
        for field, col_name in _DATE_FIELDS:
            d = _parse_date_val(row.get(field))
            if d is not None:
                dates.append((d, col_name))

        for i in range(1, len(dates)):
            if dates[i][0] < dates[i - 1][0]:
                errors.append({
                    "row_number": row["row_number"],
                    "container_no": row.get("container_no"),
                    "column": dates[i][1],
                    "value": str(dates[i][0]),
                    "reason": (
                        f"{dates[i - 1][1]}({dates[i - 1][0]})보다 "
                        f"{dates[i][1]}({dates[i][0]})이 이전 날짜입니다"
                    ),
                })

    return errors


# ─── 전체 검증 실행 ───────────────────────────────────────────────────

def run_checklist(filename: str, rows: list[dict]) -> dict:
    _refresh_holiday_set()
    errors = {
        "1": check_stage1(rows),
        "2": check_stage2(rows),
        "3": check_stage3(rows),
        "5": check_stage5(rows),
        "6": check_stage6(rows),
        "7": check_stage7(rows),
    }

    session_id = _next_session_id()
    session = {
        "id": session_id,
        "filename": filename,
        "uploaded_at": datetime.now().isoformat(),
        "total_rows": len(rows),
        "errors": errors,
    }

    _save_session(session_id, session, rows)
    return session


def run_stage4(session_id: int, container_numbers: list[str]) -> list[dict]:
    data = _load_session(session_id)
    if not data:
        return []
    rows = data.get("rows", [])
    errors = check_stage4(rows, container_numbers)
    data["session"]["errors"]["4"] = errors
    _save_session(session_id, data["session"], rows)
    return errors


# ─── 세션 저장/로드 ───────────────────────────────────────────────────

def _next_session_id() -> int:
    CHECKLIST_DIR.mkdir(parents=True, exist_ok=True)
    existing = list(CHECKLIST_DIR.glob("session_*.json"))
    if not existing:
        return 1
    ids = []
    for p in existing:
        try:
            ids.append(int(p.stem.split("_")[1]))
        except (ValueError, IndexError):
            pass
    return max(ids, default=0) + 1


def _save_session(session_id: int, session: dict, rows: list[dict]):
    CHECKLIST_DIR.mkdir(parents=True, exist_ok=True)
    path = CHECKLIST_DIR / f"session_{session_id}.json"
    data = {"session": session, "rows": rows}
    path.write_text(json.dumps(data, ensure_ascii=False, default=str), encoding="utf-8")


def _load_session(session_id: int) -> Optional[dict]:
    path = CHECKLIST_DIR / f"session_{session_id}.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def get_session(session_id: int) -> Optional[dict]:
    data = _load_session(session_id)
    if data:
        return data["session"]
    return None


def list_sessions() -> list[dict]:
    CHECKLIST_DIR.mkdir(parents=True, exist_ok=True)
    sessions = []
    for p in sorted(CHECKLIST_DIR.glob("session_*.json"), reverse=True):
        try:
            data = json.loads(p.read_text(encoding="utf-8"))
            sessions.append(data["session"])
        except Exception:
            pass
    return sessions[:20]


# ─── 리포트 엑셀 생성 ─────────────────────────────────────────────────

STAGE_NAMES = {
    "1": "필수 항목 공란 체크",
    "2": "리퍼 / DG 체크",
    "3": "공휴일 / 주말 체크",
    "4": "직상차 체크",
    "5": "운송유형 / 모드 누락 체크",
    "6": "왕복 / 편도 체크",
    "7": "일자 순서 체크",
}

_FILL_HEADER = PatternFill("solid", fgColor="1A73E8")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FILL_ERROR = PatternFill("solid", fgColor="FFF5F5")
_FILL_STAGE_HEADER = PatternFill("solid", fgColor="F3F4F6")


def generate_checklist_report(session: dict) -> bytes:
    errors = session.get("errors", {})
    wb = Workbook()

    # ── 요약 시트 ──
    ws = wb.active
    ws.title = "요약"
    ws.append(["파일명", session.get("filename", "")])
    ws.append(["업로드일시", session.get("uploaded_at", "")])
    ws.append(["총 행수", session.get("total_rows", 0)])
    ws.append([])
    ws.append(["Stage", "검증 항목", "오류 건수"])
    for cell in ws[5]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    total_errors = 0
    for stage_num in ("1", "2", "3", "4", "5", "6", "7"):
        stage_errors = errors.get(stage_num, [])
        count = len(stage_errors)
        total_errors += count
        ws.append([f"Stage {stage_num}", STAGE_NAMES.get(stage_num, ""), count])

    ws.append([])
    ws.append(["", "총 오류 건수", total_errors])
    ws[ws.max_row][2].font = Font(bold=True, color="DC2626", size=12)

    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30

    # ── 전체 오류 시트 ──
    ws2 = wb.create_sheet("전체 오류 목록")
    detail_headers = ["Stage", "검증 항목", "행번호", "컨테이너번호", "열", "현재값", "오류 사유"]
    ws2.append(detail_headers)
    for cell in ws2[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    for stage_num in ("1", "2", "3", "4", "5", "6", "7"):
        stage_errors = errors.get(stage_num, [])
        for err in stage_errors:
            ws2.append([
                f"Stage {stage_num}",
                STAGE_NAMES.get(stage_num, ""),
                err.get("row_number"),
                err.get("container_no"),
                err.get("column"),
                err.get("value", ""),
                err.get("reason"),
            ])
            ws2[ws2.max_row][0].fill = _FILL_ERROR

    widths = [12, 25, 10, 18, 20, 15, 55]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:G{ws2.max_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
