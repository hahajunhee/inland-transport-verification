import pandas as pd
import re
from io import BytesIO
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter

# ─── 컬럼 헤더 매핑 ───────────────────────────────────────────
COLUMN_MAP = {
    "D/G여부":                "dg_flag",
    "Cont.Category":          "cont_category",
    "Cont Size":              "cont_size",
    "Cont.Type":              "cont_type",
    "픽업지":                 "pickup_code",
    "픽업지명":               "pickup_name",
    "상세 ODCY":              "odcy_code",
    "상세 ODCY명":            "odcy_name",
    "ODCY도착지명":           "odcy_destination_name",
    "출하지명":               "departure_name",
    "도착지":                 "dest_code",
    "도착지명":               "dest_name",
    "출하일":                 "transport_date",
    "Contrainer No.":         "container_no",
    "C/Invoice No.":          "c_invoice_no",
    "FWO Doc.":               "fwo_doc",
    "Quantity":               "quantity",
    "Weekend / Holiday":      "weekend_holiday",
    "Mobis 운임합계(매출)":   "trkv_actual",
    "ODCY 보관료":            "storage_actual",
    "ODCY 상하차료":          "handling_actual",
    "ODCY 셔틀료":            "shuttle_actual",
    "ODCY 반입일":            "odcy_in_date",
    "ODCY 반출일":            "odcy_out_date",
}


def _safe_float(value) -> float:
    """쉼표 포함 숫자 문자열도 파싱"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _derive_container_type(row: dict) -> str:
    """
    Cont.Category (DR/RF) + D/G여부 + Cont.Type(앞2자리) → 컨테이너 유형 문자열
    예) DR + 20 + DG=N → "20드라이"
        RF + 40 + DG=Y → "40리퍼위험물"
    """
    category = str(row.get("cont_category") or "").strip().upper()
    dg_raw = str(row.get("dg_flag") or "").strip().upper()
    is_dg = dg_raw in ("Y", "TRUE", "1", "예", "X")

    # 사이즈: Cont.Type 앞 두 자리 숫자 우선
    cont_type_str = str(row.get("cont_type") or "")
    size_match = re.match(r"(\d{2})", cont_type_str)
    if size_match:
        size_num = int(size_match.group(1))
        # ISO: 22=20ft, 42=40ft, 45=45ft, 25=25ft rare
        size_map = {22: "20", 42: "40", 45: "45", 25: "25"}
        size = size_map.get(size_num, str(size_num))
    else:
        # fallback: Cont Size 필드
        cont_size = str(row.get("cont_size") or "").strip().upper()
        if "20" in cont_size:
            size = "20"
        elif "45" in cont_size:
            size = "45"
        else:
            size = "40"

    type_str = "드라이" if category == "DR" else "리퍼" if category == "RF" else category
    dg_str = "위험물" if is_dg else ""
    return f"{size}{type_str}{dg_str}"


def _parse_date(value) -> Optional[str]:
    """날짜값을 문자열로 정규화. datetime 객체, 문자열 모두 처리."""
    if value is None:
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    s = str(value).strip()
    if s in ("nan", "None", "NaT", ""):
        return None
    return s


def parse_settlement_excel(file_bytes: bytes) -> list[dict]:
    """업로드된 엑셀 파일 파싱. 헤더행 자동 탐지."""
    df = pd.read_excel(BytesIO(file_bytes), header=None, dtype=str)

    # 헤더 행 탐지: "Cont.Category" 또는 "픽업지" 포함된 행
    header_row_idx = None
    for i, row in df.iterrows():
        row_values = [str(v) for v in row.values]
        if any(col in row_values for col in ["Cont.Category", "픽업지", "D/G여부"]):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("엑셀에서 헤더 행을 찾을 수 없습니다. 컬럼명(픽업지, Cont.Category 등)이 포함된 헤더 행이 필요합니다.")

    df.columns = df.iloc[header_row_idx]
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)

    # 필요 컬럼만 추출
    available = {col: mapped for col, mapped in COLUMN_MAP.items() if col in df.columns}
    missing_required = [col for col in ["픽업지", "도착지"] if col not in available]
    if missing_required:
        raise ValueError(f"필수 컬럼 없음: {missing_required}")

    rows = []
    for idx, raw_row in df.iterrows():
        row: dict = {}
        for excel_col, internal_key in available.items():
            row[internal_key] = raw_row.get(excel_col)

        # 빈 행 스킵 (컨테이너번호 또는 픽업지가 비어있으면)
        pickup = str(row.get("pickup_code") or "").strip()
        if not pickup or pickup in ("nan", "None"):
            continue

        # 금액 파싱
        for field in ("trkv_actual", "storage_actual", "handling_actual", "shuttle_actual"):
            row[field] = _safe_float(row.get(field))

        # 수량 파싱 (없으면 1.0 기본값)
        qty = _safe_float(row.get("quantity"))
        row["quantity"] = qty if qty > 0 else 1.0

        # 컨테이너 유형 도출
        row["container_type"] = _derive_container_type(row)

        # DG 플래그: 원본값 보존 후 bool 변환
        dg_raw = str(row.get("dg_flag") or "").strip().upper()
        row["dg_raw"] = dg_raw  # TRKV 서비스에서 "X" 판단용 원본값
        row["dg_flag"] = dg_raw in ("Y", "TRUE", "1", "예", "X")

        # 날짜 정규화
        td = str(row.get("transport_date") or "").strip()
        row["transport_date"] = td if td not in ("nan", "None", "") else None

        # ODCY 반입일/반출일 파싱
        row["odcy_in_date"] = _parse_date(row.get("odcy_in_date"))
        row["odcy_out_date"] = _parse_date(row.get("odcy_out_date"))

        # 코드/이름 strip
        for field in ("pickup_code", "pickup_name", "odcy_code", "odcy_name", "odcy_destination_name", "dest_code", "dest_name", "container_no", "c_invoice_no", "fwo_doc", "departure_name"):
            val = str(row.get(field) or "").strip()
            row[field] = val if val not in ("nan", "None", "") else None

        # Weekend / Holiday: "X" 여부만 필요하므로 대문자 strip
        wh = str(row.get("weekend_holiday") or "").strip().upper()
        row["weekend_holiday"] = "X" if wh == "X" else ""

        row["row_number"] = int(idx) + header_row_idx + 2  # 실제 엑셀 행 번호
        rows.append(row)

    return rows


# ─── 결과 엑셀 생성 ───────────────────────────────────────────

# 행 상태별 색상 (웹 CSS와 동일)
FILL_OK      = PatternFill("solid", fgColor="FFFFFF")  # OK: 흰색 (웹과 동일)
FILL_DIFF    = PatternFill("solid", fgColor="FFF5F5")  # DIFF: #fff5f5
FILL_NO_RATE = PatternFill("solid", fgColor="FFFEF0")  # NO_RATE: #fffef0
FILL_SKIP    = PatternFill("solid", fgColor="F2F2F2")  # SKIP: 회색

STATUS_FILL = {
    "OK":      FILL_OK,
    "DIFF":    FILL_DIFF,
    "NO_RATE": FILL_NO_RATE,
    "SKIP":    FILL_SKIP,
}

# 섹션별 색상 (웹 CSS와 동일)
# 그룹 헤더 행 (Row 1): 진한 색
# 컬럼 헤더 행 (Row 2): 연한 색 (웹 th-xxx 배경색)
_SECTIONS = [
    # (start_col, end_col, label, group_bg, col_bg, col_font)
    (1,   5,  "기본 정보",      "374151", "E5E7EB", "374151"),
    (6,   17, "운송 구간 정보", "0F766E", "CCFBF1", "0F766E"),
    (18,  22, "TRKV",           "1A73E8", "E8F0FE", "1A73E8"),
    (23,  33, "구분값 정보",    "6B21A8", "F3E8FF", "6B21A8"),
    (34,  39, "보관료",         "1E7E34", "E6F9F0", "1E7E34"),
    (40,  43, "상하차료",       "D96C00", "FEF3E8", "D96C00"),
    (44,  47, "셔틀비용",       "7B1FA2", "F3E8FE", "7B1FA2"),
    (48,  48, "종합",           "374151", "E5E7EB", "374151"),
]

def _col_style(col_idx: int):
    """컬럼 인덱스(1-based)에 따른 (col_bg, col_font) 반환."""
    for start, end, _, _, col_bg, col_font in _SECTIONS:
        if start <= col_idx <= end:
            return col_bg, col_font
    return "FFFFFF", "000000"


def generate_results_excel(results: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "정산검증결과"

    headers = [
        "행번호", "컨테이너번호", "FWO Doc.", "C/Invoice No.", "운송일자", "픽업지코드", "픽업지명",
        "ODCY코드", "ODCY명", "도착지코드", "도착지명", "컨테이너유형", "위험물", "수량", "주말/휴일", "티어번호", "TRKV요율#",
        # TRKV
        "TRKV단가", "TRKV청구금액", "TRKV예상금액", "TRKV차이금액", "TRKV상태",
        # 구분값 정보
        "ODCY도착지명", "도착지명(원본)", "odcy터미널구분", "ODCY_위치", "도착지포트구분", "도착지터미널구분",
        "ODCY반입일", "ODCY반출일", "보관일수", "FREE반영", "보관요율#",
        # 보관료
        "보관료티어", "보관단가(일)", "보관료청구금액", "보관료예상금액", "보관료차이금액", "보관료상태",
        # 상하차료
        "상하차료청구금액", "상하차료예상금액", "상하차료차이금액", "상하차료상태",
        # 셔틀비용
        "셔틀료청구금액", "셔틀료예상금액", "셔틀료차이금액", "셔틀료상태",
        "종합상태",
    ]
    total_cols = len(headers)

    # ── Row 1: 섹션 그룹 헤더 (병합 + 진한 배경) ──────────────────
    ws.append([""] * total_cols)
    for start, end, label, group_bg, _, _ in _SECTIONS:
        cell = ws.cell(row=1, column=start, value=label)
        cell.fill = PatternFill("solid", fgColor=group_bg)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if end > start:
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1, end_column=end)
    ws.row_dimensions[1].height = 20

    # ── Row 2: 개별 컬럼 헤더 (섹션 배경색) ──────────────────────
    ws.append(headers)
    for col_idx, cell in enumerate(ws[2], 1):
        col_bg, col_font = _col_style(col_idx)
        cell.fill = PatternFill("solid", fgColor=col_bg)
        cell.font = Font(bold=True, color=col_font, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── Row 3: 출처 정보 행 ────────────────────────────────────────
    sources = [
        # 기본 정보 (1-5)
        "생성", "검증: Contrainer No.", "검증: FWO Doc.", "검증: C/Invoice No.", "검증: 출하일",
        # 운송 구간 정보 (5-16)
        "검증: 픽업지명", "요율표: PM-A", "검증: 출하지명", "요율표: DM-A",
        "검증: 상세ODCY", "검증: 도착지명", "요율표: PM-A",
        "검증: Cont.Category", "검증: Quantity", "검증: Weekend/Holiday", "요율표: 컨테이너티어", "요율표: TRKV구간",
        # TRKV (17-21)
        "요율표: TRKV구간", "검증: Mobis운임합계", "계산", "계산", "계산",
        # 구분값 정보 (22-32)
        "검증: ODCY도착지명", "검증: 도착지명",
        "요율표: OM-C", "요율표: OM-D", "요율표: PM-B", "요율표: PM-C",
        "검증: ODCY 반입일", "검증: ODCY 반출일", "계산", "계산", "요율표: 보관료",
        # 보관료 (33-38)
        "요율표: 보관료", "요율표: 보관료", "검증: ODCY 보관료", "계산", "계산", "계산",
        # 상하차료 (39-42)
        "검증: ODCY 상하차료", "계산", "계산", "계산",
        # 셔틀비용 (43-46)
        "검증: ODCY 셔틀료", "계산", "계산", "계산",
        # 종합 (47)
        "계산",
    ]
    ws.append(sources)
    for col_idx, cell in enumerate(ws[3], 1):
        col_bg, col_font = _col_style(col_idx)
        cell.fill = PatternFill("solid", fgColor=col_bg)
        cell.font = Font(size=8, color=col_font, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 16

    money_cols = {headers.index(h) + 1 for h in headers if "금액" in h}

    for r in results:
        g = (lambda k: r.get(k) if isinstance(r, dict) else getattr(r, k, None))
        row_data = [
            g("row_number"), g("container_no"), g("fwo_doc"), g("c_invoice_no"), g("transport_date"),
            g("pickup_code"), g("pickup_name"),
            g("odcy_code"), g("odcy_name"),
            g("dest_code"), g("dest_name"),
            g("container_type"), "Y" if g("dg_flag") else "N", g("quantity"), g("weekend_holiday") or "", g("tier_number"), g("trkv_rate_row"),
            g("trkv_unit_rate"), g("trkv_actual"), g("trkv_expected"), g("trkv_diff"), g("trkv_status"),
            # 구분값 정보
            g("odcy_destination_name"), g("dest_name"),
            g("odcy_terminal_type"), g("odcy_location"), g("dest_port_type"), g("dest_terminal_type"),
            g("odcy_in_date"), g("odcy_out_date"), g("storage_days"), g("billable_days"), g("storage_rate_row"),
            # 보관료
            g("storage_tier_number"), g("storage_unit_rate"), g("storage_actual"), g("storage_expected"), g("storage_diff"), g("storage_status"),
            g("handling_actual"), g("handling_expected"), g("handling_diff"), g("handling_status"),
            g("shuttle_actual"), g("shuttle_expected"), g("shuttle_diff"), g("shuttle_status"),
            g("overall_status"),
        ]
        ws.append(row_data)
        excel_row = ws.max_row

        # 행 색상 (overall_status 기준, 웹과 동일)
        row_fill = STATUS_FILL.get(g("overall_status") or "", FILL_SKIP)
        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = row_fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            if col_idx in money_cols and cell.value is not None:
                cell.number_format = '#,##0'

    # ── 컬럼 너비 자동 조정 (헤더 행 기준 포함) ───────────────────
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # 행 고정 (헤더 2행 고정)
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── FWO Charge 엑셀 생성 (DIFF → 매출인보이스 변환) ─────────────

# DIFF 발생 운임항목 → Charge Type 매핑
_CHARGE_TYPE_MAP = {
    "trkv":     "TRKV",
    "storage":  "STWV",
    "handling": "LULV",
    "shuttle":  "SHTV",
}

_FWO_HEADERS = [
    "FWO No.", "B/L No.", "Traffic Direction", "Container No.",
    "Stage Type", "Charge Type", "Currency", "Rate Amount",
    "Quantity", "Rounding Profile", "Calc. Amount",
    "Tax Code", "Tax Amount",
    "청구통화\nKRW for Billing", "Reason Code", "Reason Detail",
]


FILL_NEGATIVE = PatternFill("solid", fgColor="FFFF00")  # 노란 음영 (음수 행)


def generate_fwo_charge_excel(results: list) -> bytes:
    """DIFF 행을 FWO Charge 템플릿으로 변환한 엑셀 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "FWO Charge"

    HEADER_ROW = 9  # 헤더는 9행

    # ── 1~8행: A열에 1 입력 ──
    for _ in range(HEADER_ROW - 1):
        ws.append([1])

    # ── 9행: 헤더 ──
    header_fill = PatternFill("solid", fgColor="FFC000")
    header_font = Font(bold=True, size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(_FWO_HEADERS)
    for cell in ws[HEADER_ROW]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[HEADER_ROW].height = 30

    # 컬럼 너비
    col_widths = [16, 12, 18, 18, 12, 14, 10, 14, 10, 18, 14, 10, 12, 16, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data_align = Alignment(vertical="center")
    money_fmt = '#,##0'

    # ── DIFF 행 → 운임항목별 행 생성 (10행부터) ──
    charge_keys = [
        ("trkv",     "trkv_status",    "trkv_diff"),
        ("storage",  "storage_status", "storage_diff"),
        ("handling", "handling_status", "handling_diff"),
        ("shuttle",  "shuttle_status", "shuttle_diff"),
    ]

    for r in results:
        if r.get("overall_status") != "DIFF":
            continue

        fwo_no = r.get("fwo_doc") or ""
        container_no = r.get("container_no") or ""

        for prefix, status_key, diff_key in charge_keys:
            if r.get(status_key) != "DIFF":
                continue

            diff_val = r.get(diff_key) or 0
            if abs(diff_val) < 1:
                continue

            charge_type = _CHARGE_TYPE_MAP[prefix]
            rate_amount = diff_val
            tax_amount = round(rate_amount * 0.1)
            is_negative = rate_amount < 0
            reason_detail = "음수 금액 - 매출인보이스 엑셀 업로드 시 해당 행 삭제 필요" if is_negative else ""

            row_data = [
                fwo_no,                 # A: FWO No.
                "",                     # B: B/L No. (빈값)
                1,                      # C: Traffic Direction (고정 1)
                container_no,           # D: Container No.
                "P",                    # E: Stage Type (고정 P)
                charge_type,            # F: Charge Type
                "KRW",                  # G: Currency (고정)
                rate_amount,            # H: Rate Amount (차이금액)
                "",                     # I: Quantity
                "",                     # J: Rounding Profile
                rate_amount,            # K: Calc. Amount (= H와 동일)
                "T1",                   # L: Tax Code (고정 T1)
                tax_amount,             # M: Tax Amount (H × 0.1)
                "",                     # N: 청구통화 KRW for Billing
                "CH01",                 # O: Reason Code (고정)
                reason_detail,          # P: Reason Detail
            ]
            ws.append(row_data)
            excel_row = ws.max_row
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.alignment = data_align
                cell.font = Font(size=10)
                if col_idx in (8, 11, 13):  # H, K, M: 금액 포맷
                    cell.number_format = money_fmt
                # 음수 행: 노란 음영
                if is_negative:
                    cell.fill = FILL_NEGATIVE

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
