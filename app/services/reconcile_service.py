"""
4단계 정산대사 검증 서비스

여러 개의 정산 엑셀(1·2단계와 동일 양식)을 병합한 "합친파일" 과
MOBIS 엑셀(3단계와 동일 양식)을 키(C/Invoice No. + Container No.)로 매칭하여
각 운임 항목 합계가 일치하는지 대사한다.

매핑 (합친파일 → 모비스):
    Mobis 운임합계(매출) → 내륙운임
    ODCY 보관료          → 보관료
    ODCY 상하차료        → 상하차료(엑셀 오타 "상하자료"는 모비스 파서가 별칭 처리)
    ODCY 셔틀료          → 셔틀료
    ODCY 직반입대기료    → 대기료
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from app.services.excel_service import parse_settlement_excel
from app.services.mobis_service import (
    parse_mobis_excel, CAT_GROVE_ODCY, CAT_GROVE, CAT_MOBIS,
)

TOLERANCE = 1.0  # 원 단위 허용 오차

# (라벨, 합친파일 필드, 모비스 cost 키, 합친파일 표시 열제목, 모비스 표시 열제목)
ITEM_MAP = [
    ("내륙운임", "trkv_actual",    "내륙운임", "Mobis 운임합계(매출)", "내륙운임"),
    ("보관료",   "storage_actual", "보관료",   "ODCY 보관료",          "보관료"),
    ("상하차료", "handling_actual", "상하차료", "ODCY 상하차료",        "상하차료"),
    ("셔틀료",   "shuttle_actual", "셔틀료",   "ODCY 셔틀료",          "셔틀료"),
    ("대기료",   "waiting_actual", "대기료",   "ODCY 직반입대기료",    "대기료"),
]
ITEM_LABELS = [m[0] for m in ITEM_MAP]

# 구분 필터 옵션
GUBUN_OPTIONS = ["MOBIS 산출", "GROVE 전송", "GROVE ODCY 전송", "ALL"]


def _key(c_inv, cont) -> tuple:
    return ((c_inv or "").strip(), (cont or "").strip())


def parse_merged_settlement(files: list[bytes]) -> dict:
    """여러 정산 엑셀을 파싱·병합 후 키별 합계.
    반환: {"agg": {(c_inv, cont): {field: sum, "_count": n}}, "total_rows": int, "per_file": [n,...]}
    """
    all_rows = []
    per_file = []
    for fb in files:
        rows = parse_settlement_excel(fb)
        per_file.append(len(rows))
        all_rows.extend(rows)

    agg: dict[tuple, dict] = {}
    for r in all_rows:
        key = _key(r.get("c_invoice_no"), r.get("container_no"))
        if key == ("", ""):
            continue
        d = agg.get(key)
        if d is None:
            d = {f: 0.0 for _, f, _, _, _ in ITEM_MAP}
            d["_count"] = 0
            agg[key] = d
        d["_count"] += 1
        for _, field, _, _, _ in ITEM_MAP:
            d[field] += float(r.get(field) or 0.0)

    return {"agg": agg, "total_rows": len(all_rows), "per_file": per_file}


def aggregate_mobis(parsed: dict, gubun_filter: str = "MOBIS 산출") -> dict:
    """MOBIS 파싱결과를 키별 합계. gubun_filter == 'ALL' 이면 구분 무시 전체 합산."""
    rows = parsed.get("rows", [])
    agg: dict[tuple, dict] = {}
    for r in rows:
        if gubun_filter and gubun_filter != "ALL" and r.get("gubun") != gubun_filter:
            continue
        key = _key(r.get("c_inv_no"), r.get("container_no"))
        if key == ("", ""):
            continue
        d = agg.get(key)
        if d is None:
            d = {label: 0.0 for label, _, _, _, _ in ITEM_MAP}
            d["_count"] = 0
            agg[key] = d
        d["_count"] += 1
        costs = r.get("costs") or {}
        for label, _, mobis_key, _, _ in ITEM_MAP:
            d[label] += float(costs.get(mobis_key) or 0.0)
    return agg


# 상태 정렬 우선순위 (고아행 먼저 → 금액불일치 → 정상)
_STATUS_ORDER = {
    "합친파일에 없음": 0,
    "모비스에 없음": 1,
    "금액불일치": 2,
    "정상": 3,
}


def _compare(s_agg: dict, m_agg: dict) -> tuple[list, dict]:
    """키 합계 두 dict 를 비교하여 결과 행 리스트 + 카운트 반환."""
    results = []
    cnt = {"ok": 0, "mismatch": 0, "settle_only": 0, "mobis_only": 0}

    for key in set(s_agg) | set(m_agg):
        c_inv, cont = key
        in_s = key in s_agg
        in_m = key in m_agg
        items = []
        reasons = []

        if in_s and in_m:
            s = s_agg[key]
            m = m_agg[key]
            is_error = False
            s_total = m_total = 0.0
            for label, field, _mk, _sc, _mc in ITEM_MAP:
                sv = float(s.get(field, 0.0))
                mv = float(m.get(label, 0.0))
                diff = sv - mv
                match = abs(diff) < TOLERANCE
                if not match:
                    is_error = True
                    reasons.append(
                        f"{label} 불일치: 합친 {sv:,.0f} ≠ 모비스 {mv:,.0f} (차이 {diff:+,.0f})"
                    )
                items.append({"label": label, "settlement": sv, "mobis": mv,
                              "diff": diff, "match": match})
                s_total += sv
                m_total += mv
            status = "금액불일치" if is_error else "정상"
            total_diff = s_total - m_total
            cnt["mismatch" if is_error else "ok"] += 1
            settlement_count = s.get("_count", 0)
            mobis_count = m.get("_count", 0)

        elif in_s:  # 합친파일에만 존재 → 모비스에 없음
            s = s_agg[key]
            is_error = True
            status = "모비스에 없음"
            s_total = 0.0
            for label, field, _mk, _sc, _mc in ITEM_MAP:
                sv = float(s.get(field, 0.0))
                items.append({"label": label, "settlement": sv, "mobis": None,
                              "diff": None, "match": False})
                s_total += sv
            m_total = None
            total_diff = None
            reasons.append("모비스 파일에 동일 키(C/INV+컨테이너)가 없습니다.")
            cnt["settle_only"] += 1
            settlement_count = s.get("_count", 0)
            mobis_count = 0

        else:  # 모비스에만 존재 → 합친파일에 없음
            m = m_agg[key]
            is_error = True
            status = "합친파일에 없음"
            m_total = 0.0
            for label, field, _mk, _sc, _mc in ITEM_MAP:
                mv = float(m.get(label, 0.0))
                items.append({"label": label, "settlement": None, "mobis": mv,
                              "diff": None, "match": False})
                m_total += mv
            s_total = None
            total_diff = None
            reasons.append("합친(정산) 파일에 동일 키(C/Invoice+Container)가 없습니다.")
            cnt["mobis_only"] += 1
            settlement_count = 0
            mobis_count = m.get("_count", 0)

        results.append({
            "container_no": cont,
            "c_inv_no": c_inv,
            "status": status,
            "is_error": is_error,
            "reasons": reasons,
            "items": items,
            "settlement_total": s_total,
            "mobis_total": m_total,
            "total_diff": total_diff,
            "settlement_count": settlement_count,
            "mobis_count": mobis_count,
        })

    results.sort(key=lambda r: (_STATUS_ORDER.get(r["status"], 9),
                                r["container_no"], r["c_inv_no"]))
    return results, cnt


def run_reconciliation(settlement_files: list[bytes], mobis_bytes: bytes,
                       gubun_filter: str = "MOBIS 산출",
                       mobis_filename: str = "결과") -> dict:
    """4단계 정산대사 실행."""
    merged = parse_merged_settlement(settlement_files)
    mobis_parsed = parse_mobis_excel(mobis_bytes)
    m_agg = aggregate_mobis(mobis_parsed, gubun_filter)

    results, cnt = _compare(merged["agg"], m_agg)
    error_count = cnt["mismatch"] + cnt["settle_only"] + cnt["mobis_only"]

    return {
        "filename": mobis_filename,
        "gubun_filter": gubun_filter,
        "settlement_files": len(settlement_files),
        "settlement_file_rows": merged["per_file"],
        "settlement_rows": merged["total_rows"],
        "settlement_keys": len(merged["agg"]),
        "mobis_keys": len(m_agg),
        "total_keys": len(results),
        "ok_count": cnt["ok"],
        "mismatch_count": cnt["mismatch"],
        "settle_only_count": cnt["settle_only"],
        "mobis_only_count": cnt["mobis_only"],
        "error_count": error_count,
        "item_labels": ITEM_LABELS,
        "results": results,
    }


# ─── 엑셀 리포트 ────────────────────────────────────────────────────────

def generate_reconcile_report(verification: dict) -> bytes:
    """4단계 대사 결과 엑셀 생성. 웹 화면과 동일한 컬럼 구조."""
    wb = Workbook()
    ws = wb.active
    ws.title = "정산대사결과"

    FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")
    FILL_OK = PatternFill("solid", fgColor="FFFFFF")
    FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
    FILL_HEADER_MAIN = PatternFill("solid", fgColor="374151")
    FILL_HEADER_SETTLE = PatternFill("solid", fgColor="1A73E8")  # 합친파일
    FILL_HEADER_MOBIS = PatternFill("solid", fgColor="0F766E")   # 모비스
    FILL_HEADER_DIFF = PatternFill("solid", fgColor="B45309")    # 차이
    FILL_HEADER_TOTAL = PatternFill("solid", fgColor="6B21A8")   # 합계
    FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
    FONT_ERROR = Font(color="9C0006", size=10)
    FONT_NORMAL = Font(size=10)
    money_fmt = '#,##0'

    labels = verification.get("item_labels", ITEM_LABELS)
    n_items = len(labels)
    # 컬럼 수: 기본 4 + (항목 n×3) + 합계 3
    base_cols = 4
    total_cols = base_cols + n_items * 3 + 3

    # ── Row 1: 그룹 헤더 ──────────────────────────────────────
    ws.append([""] * total_cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=base_cols)
    c = ws.cell(row=1, column=1, value="기본 정보")
    c.fill = FILL_HEADER_MAIN; c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")

    col = base_cols + 1
    group_fill_cycle = [FILL_HEADER_SETTLE, FILL_HEADER_MOBIS, FILL_HEADER_DIFF]
    for label in labels + ["합계"]:
        is_total = (label == "합계")
        gfill = FILL_HEADER_TOTAL if is_total else FILL_HEADER_MAIN
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
        c = ws.cell(row=1, column=col, value=label)
        c.fill = gfill; c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += 3
    ws.row_dimensions[1].height = 22

    # ── Row 2: 세부 헤더 ──────────────────────────────────────
    headers2 = ["컨테이너 번호", "C/INV 번호", "결과", "사유"]
    for _ in labels + ["합계"]:
        headers2 += ["합친파일", "모비스", "차이"]
    ws.append(headers2)
    # 색상
    header_fills = [FILL_HEADER_MAIN] * base_cols
    for _ in labels:
        header_fills += [FILL_HEADER_SETTLE, FILL_HEADER_MOBIS, FILL_HEADER_DIFF]
    header_fills += [FILL_HEADER_TOTAL, FILL_HEADER_TOTAL, FILL_HEADER_TOTAL]
    for ci, cell in enumerate(ws[2], 1):
        cell.fill = header_fills[ci - 1] if ci - 1 < len(header_fills) else FILL_HEADER_MAIN
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # ── 데이터 행 ─────────────────────────────────────────────
    for r in verification.get("results", []):
        is_err = r["is_error"]
        row_data = [
            r["container_no"], r["c_inv_no"], r["status"],
            " / ".join(r.get("reasons", [])),
        ]
        items = {it["label"]: it for it in r.get("items", [])}
        for label in labels:
            it = items.get(label, {})
            sv = it.get("settlement")
            mv = it.get("mobis")
            dv = it.get("diff")
            row_data += [
                sv if sv is not None else "",
                mv if mv is not None else "",
                dv if dv is not None else "",
            ]
        row_data += [
            r.get("settlement_total") if r.get("settlement_total") is not None else "",
            r.get("mobis_total") if r.get("mobis_total") is not None else "",
            r.get("total_diff") if r.get("total_diff") is not None else "",
        ]
        ws.append(row_data)
        er = ws.max_row
        base_fill = FILL_ERROR if is_err else FILL_OK
        base_font = FONT_ERROR if is_err else FONT_NORMAL
        for ci in range(1, total_cols + 1):
            cell = ws.cell(row=er, column=ci)
            cell.fill = base_fill
            cell.font = base_font
            cell.alignment = Alignment(vertical="center")
            if ci > base_cols and isinstance(cell.value, (int, float)):
                cell.number_format = money_fmt

        # 항목별 불일치 셀 노란 음영 (차이 컬럼 + 양쪽 값)
        for i, label in enumerate(labels):
            it = items.get(label, {})
            if not it.get("match", True):
                base = base_cols + i * 3  # 1-based 직전
                for off in (1, 2, 3):
                    ws.cell(row=er, column=base + off).fill = FILL_YELLOW

    # ── 열 너비 ──────────────────────────────────────────────
    widths = [18, 18, 14, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(n_items * 3 + 3):
        ws.column_dimensions[get_column_letter(base_cols + 1 + i)].width = 14
    ws.freeze_panes = "E3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
