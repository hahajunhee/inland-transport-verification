"""
3단계 모비스 검증 서비스
엑셀 업로드 → 컨테이너번호+C/INV번호 중복제거 → 구분별 AE:AI 합계 비교 → 오류 판정
"""
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ─── 열제목 매핑 ────────────────────────────────────────────────────────
_REQUIRED_HEADERS = ["컨테이너 번호", "C/INV\n번호", "구분"]
_COST_HEADERS = ["내륙운임", "보관료", "상하차료", "셔틀료", "대기료"]
_EXTRA_HEADERS = ["적요", "출발지", "작업지", "경유지"]

# 구분 카테고리
CAT_GROVE_ODCY = "GROVE ODCY 전송"
CAT_GROVE = "GROVE 전송"
CAT_MOBIS = "MOBIS 산출"
_CATEGORIES = [CAT_GROVE_ODCY, CAT_GROVE, CAT_MOBIS]


def _safe_float(value) -> float:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _safe_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return "" if s in ("nan", "None", "NaT") else s


def _find_header_columns(df: pd.DataFrame) -> dict:
    """
    1:2행 병합 헤더에서 필요 컬럼 위치를 찾는다.
    2-pass: 정확 매칭 우선, 미발견 헤더만 부분 매칭.
    반환: {header_name: column_index}
    """
    import re

    ALL_HEADERS = _REQUIRED_HEADERS + _COST_HEADERS + _EXTRA_HEADERS
    nrows = min(5, len(df))
    col_map = {}
    dest_cols = []  # "도착지" 중복 처리용

    def _norm(text: str) -> str:
        """모든 공백(\n, \r, \t, 스페이스 등) 제거."""
        return re.sub(r'\s+', '', text)

    # 각 컬럼의 후보 텍스트 수집
    col_candidates = []
    for ci in range(len(df.columns)):
        candidates = []
        for ri in range(nrows):
            val = str(df.iloc[ri, ci]).strip()
            if val and val != "nan":
                candidates.append(val)
        col_candidates.append(candidates)

        # "도착지" 개별 수집 (중복 가능 — 구간/ODCY)
        for val in candidates:
            if _norm(val) == "도착지":
                dest_cols.append(ci)
                break

    # Pass 1: 개별 후보에서 정확 매칭 (우선)
    for ci, candidates in enumerate(col_candidates):
        for header in ALL_HEADERS:
            if header in col_map:
                continue
            norm_header = _norm(header)
            for val in candidates:
                if _norm(val) == norm_header:
                    col_map[header] = ci
                    break

    # Pass 2: combined 텍스트에서 부분 매칭 (미발견 헤더만)
    for ci, candidates in enumerate(col_candidates):
        norm_combined = _norm("".join(candidates))
        for header in ALL_HEADERS:
            if header in col_map:
                continue
            if _norm(header) in norm_combined:
                col_map[header] = ci

    # "도착지" 중복 처리: 첫번째=구간 도착지, 두번째=ODCY 도착지
    if len(dest_cols) >= 2:
        col_map["도착지"] = dest_cols[0]
        col_map["ODCY도착지"] = dest_cols[1]
    elif len(dest_cols) == 1:
        col_map["도착지"] = dest_cols[0]

    return col_map


def _find_data_start(df: pd.DataFrame, col_map: dict) -> int:
    """헤더 행 이후 실제 데이터 시작 행 인덱스 반환."""
    gubun_col = col_map.get("구분")
    container_col = col_map.get("컨테이너 번호")

    if gubun_col is None:
        return 2

    for ri in range(len(df)):
        val = str(df.iloc[ri, gubun_col]).strip()
        if val in _CATEGORIES:
            return ri
        if container_col is not None:
            cont_val = str(df.iloc[ri, container_col]).strip()
            if len(cont_val) >= 10 and cont_val[:4].isalpha():
                return ri

    return 2


def parse_mobis_excel(file_bytes: bytes) -> dict:
    """모비스 검증 엑셀 파싱."""
    df = pd.read_excel(BytesIO(file_bytes), header=None, dtype=str)
    col_map = _find_header_columns(df)

    # 필수 컬럼 확인
    missing = [h for h in _REQUIRED_HEADERS if h not in col_map]
    if missing:
        raise ValueError(f"필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}")

    found_costs = [h for h in _COST_HEADERS if h in col_map]
    if not found_costs:
        raise ValueError("비용 컬럼(내륙운임, 보관료, 상하차료, 셔틀료, 대기료)을 찾을 수 없습니다.")

    data_start = _find_data_start(df, col_map)

    # 구간 컬럼 키 목록
    route_keys = ["출발지", "작업지", "경유지", "도착지", "ODCY도착지"]

    rows = []
    for ri in range(data_start, len(df)):
        container_no = _safe_str(df.iloc[ri, col_map["컨테이너 번호"]])
        c_inv = _safe_str(df.iloc[ri, col_map["C/INV\n번호"]])
        gubun = _safe_str(df.iloc[ri, col_map["구분"]])

        if not container_no and not c_inv:
            continue

        # 비용 항목별 값
        costs = {}
        for h in _COST_HEADERS:
            costs[h] = _safe_float(df.iloc[ri, col_map[h]]) if h in col_map else 0.0

        # 적요
        jukyo = ""
        if "적요" in col_map:
            jukyo = _safe_str(df.iloc[ri, col_map["적요"]])

        # 구간 정보
        route = {}
        for rk in route_keys:
            if rk in col_map:
                route[rk] = _safe_str(df.iloc[ri, col_map[rk]])
            else:
                route[rk] = ""

        rows.append({
            "container_no": container_no,
            "c_inv_no": c_inv,
            "gubun": gubun,
            "costs": costs,
            "cost_sum": sum(costs.values()),
            "jukyo": jukyo,
            "route": route,
            "row_number": ri + 1,
        })

    return {
        "rows": rows,
        "cost_headers": found_costs,
    }


_REF_FIELDS = ["pickup_name", "departure_name", "odcy_code", "dest_port_type", "odcy_destination_name"]
_REF_EMPTY = {f: "존재X" for f in _REF_FIELDS}

_EXPECTED_FIELDS = ["trkv_expected", "storage_expected", "handling_expected", "shuttle_expected"]
_EXPECTED_EMPTY = {f: "존재X" for f in _EXPECTED_FIELDS}
_EXPECTED_TO_MOBIS = {
    "trkv_expected": "내륙운임",
    "storage_expected": "보관료",
    "handling_expected": "상하차료",
    "shuttle_expected": "셔틀료",
}


def build_ref_lookup(session_id: int) -> dict:
    """2단계 검증결과에서 (c_invoice_no, container_no) → 참조 데이터 dict 빌드."""
    from app import data_store
    results = data_store.load_results(session_id)
    lookup = {}
    for r in results:
        c_inv = (r.get("c_invoice_no") or "").strip()
        cont = (r.get("container_no") or "").strip()
        if not c_inv and not cont:
            continue
        key = (c_inv, cont)
        if key not in lookup:
            data = {f: (r.get(f) or "") for f in _REF_FIELDS}
            for f in _EXPECTED_FIELDS:
                data[f] = r.get(f)
            lookup[key] = data
    return lookup


def run_mobis_verification(filename: str, parsed: dict, ref_lookup: dict = None) -> dict:
    """
    모비스 검증 실행.
    오류 조건:
      1) GROVE 전송 합계 ≠ MOBIS 산출 합계
      2) GROVE ODCY 전송 행이 존재
    ref_lookup: build_ref_lookup()으로 생성한 2단계 검증결과 참조 dict (선택).
    """
    rows = parsed["rows"]
    cost_headers = parsed["cost_headers"]

    # (container_no, c_inv_no) 별로 그룹핑
    groups: dict[tuple, list] = {}
    group_order = []
    for row in rows:
        key = (row["container_no"], row["c_inv_no"])
        if key not in groups:
            groups[key] = []
            group_order.append(key)
        groups[key].append(row)

    results = []
    error_count = 0
    ok_count = 0

    for key in group_order:
        container_no, c_inv_no = key
        group_rows = groups[key]

        # 구분별 합계 + 항목별 세부값 계산
        cat_sums = {}
        cat_details = {}
        has_grove_odcy = False
        for cat in _CATEGORIES:
            matching = [r for r in group_rows if r["gubun"] == cat]
            if matching:
                if cat == CAT_GROVE_ODCY:
                    has_grove_odcy = True
                total = sum(r["cost_sum"] for r in matching)
                detail = {}
                for h in cost_headers:
                    detail[h] = sum(r["costs"].get(h, 0.0) for r in matching)
                cat_sums[cat] = total
                cat_details[cat] = detail
            else:
                cat_sums[cat] = None
                cat_details[cat] = None

        grove_odcy_sum = cat_sums.get(CAT_GROVE_ODCY) or 0.0
        grove_sum = cat_sums.get(CAT_GROVE) if cat_sums.get(CAT_GROVE) is not None else 0.0
        mobis_sum = cat_sums.get(CAT_MOBIS) if cat_sums.get(CAT_MOBIS) is not None else 0.0

        # ── 오류 판정 ──
        # 1) GROVE 전송 합계 ≠ MOBIS 산출 합계
        # 2) GROVE ODCY 전송 행이 존재
        error_reasons = []
        if abs(grove_sum - mobis_sum) >= 1:
            error_reasons.append(f"GROVE 전송({grove_sum:,.0f}) ≠ MOBIS 산출({mobis_sum:,.0f})")
        if has_grove_odcy:
            error_reasons.append("GROVE ODCY 전송 존재")
        is_error = len(error_reasons) > 0

        if is_error:
            error_count += 1
        else:
            ok_count += 1

        # 적요 수집 (GROVE ODCY 전송 행 우선, 없으면 아무 행)
        jukyo = ""
        grove_odcy_rows = [r for r in group_rows if r["gubun"] == CAT_GROVE_ODCY]
        if grove_odcy_rows:
            jukyo = grove_odcy_rows[0].get("jukyo", "")
        if not jukyo:
            for r in group_rows:
                if r.get("jukyo"):
                    jukyo = r["jukyo"]
                    break

        # 구간 정보 (MOBIS 산출 행에서 가져오기)
        route_info = {"출발지": "", "작업지": "", "경유지": "", "도착지": "", "ODCY도착지": ""}
        mobis_rows = [r for r in group_rows if r["gubun"] == CAT_MOBIS]
        if mobis_rows:
            route_info = mobis_rows[0].get("route", route_info)

        # 정산검증 참조 (2단계 검증결과에서 키 매칭)
        ref = dict(_REF_EMPTY) if ref_lookup is not None else {f: "" for f in _REF_FIELDS}
        expected = dict(_EXPECTED_EMPTY) if ref_lookup is not None else {f: "" for f in _EXPECTED_FIELDS}
        if ref_lookup is not None:
            ref_key = (c_inv_no, container_no)
            if ref_key in ref_lookup:
                ref_data = ref_lookup[ref_key]
                ref = {f: ref_data.get(f, "") for f in _REF_FIELDS}
                for f in _EXPECTED_FIELDS:
                    expected[f] = ref_data.get(f)

        # 세부 비교: GROVE 전송 vs MOBIS 산출 항목별 차이 (오류행용)
        detail_diffs = {}
        grove_detail = cat_details.get(CAT_GROVE) or {}
        mobis_detail = cat_details.get(CAT_MOBIS) or {}
        for h in cost_headers:
            gv = grove_detail.get(h, 0.0)
            mv = mobis_detail.get(h, 0.0)
            detail_diffs[h] = abs(gv - mv) >= 1

        # 예상금액 vs MOBIS 산출 비교
        expected_diffs = {}
        for exp_key, mobis_key in _EXPECTED_TO_MOBIS.items():
            exp_val = expected.get(exp_key)
            if exp_val is None or exp_val == "" or exp_val == "존재X":
                expected_diffs[exp_key] = False
            else:
                mobis_val = mobis_detail.get(mobis_key, 0.0)
                expected_diffs[exp_key] = abs(_safe_float(exp_val) - mobis_val) >= 1

        result = {
            "container_no": container_no,
            "c_inv_no": c_inv_no,
            "is_error": is_error,
            "status": "오류" if is_error else "정상",
            "error_reasons": error_reasons,
            "jukyo": jukyo,
            "grove_odcy_sum": grove_odcy_sum,
            "grove_sum": grove_sum,
            "mobis_sum": mobis_sum,
            "grove_detail": grove_detail,
            "mobis_detail": mobis_detail,
            "detail_diffs": detail_diffs,
            "route": route_info,
            "ref": ref,
            "expected": expected,
            "expected_diffs": expected_diffs,
        }
        results.append(result)

    return {
        "filename": filename,
        "total_groups": len(results),
        "error_count": error_count,
        "ok_count": ok_count,
        "cost_headers": cost_headers,
        "results": results,
    }


def generate_mobis_report(verification: dict) -> bytes:
    """모비스 검증 결과 엑셀 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "모비스검증결과"

    cost_headers = _COST_HEADERS  # 항상 5개 비용 항목 사용

    FILL_ERROR = PatternFill("solid", fgColor="FFC7CE")
    FILL_OK = PatternFill("solid", fgColor="FFFFFF")
    FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
    FILL_HEADER_MAIN = PatternFill("solid", fgColor="374151")
    FILL_HEADER_GROVE = PatternFill("solid", fgColor="1A73E8")
    FILL_HEADER_MOBIS = PatternFill("solid", fgColor="0F766E")
    FILL_HEADER_ROUTE = PatternFill("solid", fgColor="6B21A8")

    FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
    FONT_ERROR = Font(color="9C0006", size=10)
    FONT_NORMAL = Font(size=10)

    nc = len(cost_headers)  # 비용 항목 수
    _REF_LABELS = ["픽업지명", "출하지명", "ODCY코드", "도착지포트구분", "ODCY도착지명"]
    _ROUTE_LABELS = ["출발지", "작업지", "경유지", "도착지", "ODCY도착지"]
    FILL_HEADER_REF = PatternFill("solid", fgColor="B45309")
    FILL_HEADER_EXPECTED = PatternFill("solid", fgColor="15803D")
    _EXPECTED_LABELS = ["TRKV예상금액", "보관료예상금액", "상하차료예상금액", "셔틀료예상금액"]

    # ── Row 1: 그룹 헤더 (병합) ─────────────────────────────
    total_cols = 9 + nc * 2 + 4 + 5 + 5  # 기본+GROVE+MOBIS+정산예상+정산참조+구간
    ws.append([""] * total_cols)
    # 기본 정보 (1~9)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    c = ws.cell(row=1, column=1, value="기본 정보")
    c.fill = FILL_HEADER_MAIN; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    # GROVE 전송 세부 (10 ~ 9+nc)
    gs = 10; ge = 9 + nc
    ws.merge_cells(start_row=1, start_column=gs, end_row=1, end_column=ge)
    c = ws.cell(row=1, column=gs, value="GROVE 전송 세부")
    c.fill = FILL_HEADER_GROVE; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    # MOBIS 산출 세부 (ge+1 ~ ge+nc)
    ms = ge + 1; me = ge + nc
    ws.merge_cells(start_row=1, start_column=ms, end_row=1, end_column=me)
    c = ws.cell(row=1, column=ms, value="MOBIS 산출 세부")
    c.fill = FILL_HEADER_MOBIS; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    # 정산 예상금액 (me+1 ~ me+4)
    es = me + 1; ee = me + 4
    ws.merge_cells(start_row=1, start_column=es, end_row=1, end_column=ee)
    c = ws.cell(row=1, column=es, value="정산 예상금액")
    c.fill = FILL_HEADER_EXPECTED; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    # 정산검증 참조 (ee+1 ~ ee+5)
    refs = ee + 1; refe = ee + 5
    ws.merge_cells(start_row=1, start_column=refs, end_row=1, end_column=refe)
    c = ws.cell(row=1, column=refs, value="정산검증 참조")
    c.fill = FILL_HEADER_REF; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    # 구간 (refe+1 ~ refe+5)
    rts = refe + 1; rte = refe + 5
    ws.merge_cells(start_row=1, start_column=rts, end_row=1, end_column=rte)
    c = ws.cell(row=1, column=rts, value="구간 (MOBIS 산출)")
    c.fill = FILL_HEADER_ROUTE; c.font = FONT_HEADER; c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: 개별 컬럼 헤더 ───────────────────────────────
    headers2 = [
        "#", "컨테이너 번호", "C/INV 번호",
        "결과", "오류사유", "오류사유_모비스",
        "GROVE ODCY\n전송 합계", "GROVE 전송\n합계", "MOBIS 산출\n합계",
    ]
    headers2 += [f"G_{h}" for h in cost_headers]
    headers2 += [f"M_{h}" for h in cost_headers]
    headers2 += _EXPECTED_LABELS
    headers2 += _REF_LABELS
    headers2 += _ROUTE_LABELS

    ws.append(headers2)
    header_fills = (
        [FILL_HEADER_MAIN] * 9
        + [FILL_HEADER_GROVE] * nc
        + [FILL_HEADER_MOBIS] * nc
        + [FILL_HEADER_EXPECTED] * 4
        + [FILL_HEADER_REF] * 5
        + [FILL_HEADER_ROUTE] * 5
    )
    for ci, cell in enumerate(ws[2], 1):
        cell.fill = header_fills[ci - 1] if ci - 1 < len(header_fills) else FILL_HEADER_MAIN
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    money_fmt = '#,##0'

    for idx, r in enumerate(verification["results"], 1):
        is_err = r["is_error"]
        gd = r.get("grove_detail") or {}
        md = r.get("mobis_detail") or {}
        dd = r.get("detail_diffs") or {}
        rf = r.get("ref") or {}
        rt = r.get("route") or {}

        row_data = [
            idx,
            r["container_no"], r["c_inv_no"],
            r["status"],
            " / ".join(r["error_reasons"]) if r["error_reasons"] else "",
            r.get("jukyo", ""),
            r["grove_odcy_sum"], r["grove_sum"], r["mobis_sum"],
        ]
        # GROVE 세부
        for h in cost_headers:
            row_data.append(gd.get(h, 0.0) if is_err else "")
        # MOBIS 세부
        for h in cost_headers:
            row_data.append(md.get(h, 0.0) if is_err else "")
        # 정산 예상금액
        exp = r.get("expected") or {}
        for f in _EXPECTED_FIELDS:
            val = exp.get(f)
            if val is None or val == "" or val == "존재X":
                row_data.append(val if val == "존재X" else "")
            else:
                row_data.append(_safe_float(val))
        # 정산검증 참조
        row_data += [rf.get(f, "") for f in _REF_FIELDS]
        # 구간
        row_data += [rt.get("출발지", ""), rt.get("작업지", ""), rt.get("경유지", ""),
                     rt.get("도착지", ""), rt.get("ODCY도착지", "")]

        ws.append(row_data)
        excel_row = ws.max_row
        base_fill = FILL_ERROR if is_err else FILL_OK
        base_font = FONT_ERROR if is_err else FONT_NORMAL

        for ci in range(1, len(row_data) + 1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.fill = base_fill
            cell.font = base_font
            cell.alignment = Alignment(vertical="center")
            if ci in (7, 8, 9):
                cell.number_format = money_fmt

        # 세부 비용 금액 포맷 + 노란 음영
        if is_err:
            for i, h in enumerate(cost_headers):
                # GROVE 세부 셀
                g_ci = 10 + i
                cell_g = ws.cell(row=excel_row, column=g_ci)
                cell_g.number_format = money_fmt
                # MOBIS 세부 셀
                m_ci = 10 + nc + i
                cell_m = ws.cell(row=excel_row, column=m_ci)
                cell_m.number_format = money_fmt
                # 차이나는 항목 → 노란 음영
                if dd.get(h, False):
                    cell_g.fill = FILL_YELLOW
                    cell_m.fill = FILL_YELLOW

        # 정산 예상금액 포맷 + 노란 음영
        ed = r.get("expected_diffs") or {}
        for i, f in enumerate(_EXPECTED_FIELDS):
            exp_ci = 10 + nc * 2 + i  # GROVE(nc) + MOBIS(nc) 이후
            cell_e = ws.cell(row=excel_row, column=exp_ci)
            val = (r.get("expected") or {}).get(f)
            if val is not None and val != "" and val != "존재X":
                cell_e.number_format = money_fmt
            if ed.get(f, False):
                cell_e.fill = FILL_YELLOW

    # ── 열 너비 ─────────────────────────────────────────────
    base_widths = [5, 18, 18, 8, 24, 24, 16, 16, 16]
    for i, w in enumerate(base_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(nc * 2):
        ws.column_dimensions[get_column_letter(10 + i)].width = 14
    # 정산 예상금액 (4열)
    exp_start = 10 + nc * 2
    for i in range(4):
        ws.column_dimensions[get_column_letter(exp_start + i)].width = 14
    ref_start = exp_start + 4
    for i in range(5):
        ws.column_dimensions[get_column_letter(ref_start + i)].width = 16
    route_start = ref_start + 5
    for i in range(5):
        ws.column_dimensions[get_column_letter(route_start + i)].width = 12

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
